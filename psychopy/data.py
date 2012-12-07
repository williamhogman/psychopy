# -*- coding: utf-8 -*-
"""Routines for handling data structures and analysis"""
# Part of the PsychoPy library
# Copyright (C) 2012 Jonathan Peirce
# Distributed under the terms of the GNU General Public License (GPL).

from psychopy import misc, gui, logging
import psychopy
import cPickle, string, sys, platform, os, time, copy, csv
import numpy
from scipy import optimize, special
from matplotlib import mlab    #used for importing csv files
from contrib.quest import *    #used for QuestHandler
import inspect #so that Handlers can find the script that called them
import codecs, locale
import weakref
import re

try:
    import openpyxl
    from openpyxl.cell import get_column_letter
    from openpyxl.reader.excel import load_workbook
    haveOpenpyxl=True
except:
    haveOpenpyxl=False

_experiments=weakref.WeakValueDictionary()
_nonalphanumeric_re = re.compile(r'\W') # will match all bad var name chars

class ExperimentHandler(object):
    """A container class for keeping track of multiple loops/handlers

    Useful for generating a single data file from an experiment with many
    different loops (e.g. interleaved staircases or loops within loops

    :usage:

        exp = data.ExperimentHandler(name="Face Preference",version='0.1.0')

    """
    def __init__(self,
                name='',
                version='',
                extraInfo=None,
                runtimeInfo=None,
                originPath=None,
                savePickle=True,
                saveWideText=True,
                dataFileName=''):
        """
        :parameters:

            name : a string or unicode
                As a useful identifier later

            version : usually a string (e.g. '1.1.0')
                To keep track of which version of the experiment was run

            extraInfo : a dictionary
                Containing useful information about this run
                (e.g. {'participant':'jwp','gender':'m','orientation':90} )

            runtimeInfo : :class:`psychopy.info.RunTimeInfo`
                Containining information about the system as detected at runtime

            originPath : string or unicode
                The path and filename of the originating script/experiment
                If not provided this will be determined as the path of the
                calling script.

            dataFilename : string
                This is defined in advance and the file will be saved at any
                point that the handler is removed or discarded (unless .abort()
                had been called in advance).
                The handler will attempt to populate the file even in the
                event of a (not too serious) crash!

        """
        self.loops=[]
        self.loopsUnfinished=[]
        self.name=name
        self.version=version
        self.runtimeInfo=runtimeInfo
        if extraInfo==None:
            self.extraInfo = {}
        else:
            self.extraInfo=extraInfo
        self.originPath=originPath
        self.savePickle=savePickle
        self.saveWideText=saveWideText
        self.dataFileName=dataFileName
        self.thisEntry = {}
        self.entries=[]#chronological list of entries
        self._paramNamesSoFar=[]
        self.dataNames=[]#names of all the data (eg. resp.keys)
        if dataFileName in ['', None]:
            logging.warning('ExperimentHandler created with no dataFileName parameter. No data will be saved in the event of a crash')
    def __del__(self):
        if self.dataFileName not in ['', None]:
            logging.debug('Saving data for %s ExperimentHandler' %self.name)
            if self.savePickle==True:
                self.saveAsPickle(self.dataFileName)
            if self.saveWideText==True:
                self.saveAsWideText(self.dataFileName+'.csv', delim=',')
    def addLoop(self, loopHandler):
        """Add a loop such as a `~psychopy.data.TrialHandler` or `~psychopy.data.StairHandler`
        Data from this loop will be included in the resulting data files.
        """
        self.loops.append(loopHandler)
        self.loopsUnfinished.append(loopHandler)
        #keep the loop updated that is now owned
        loopHandler.setExp(self)
    def loopEnded(self, loopHandler):
        """Informs the experiment handler that the loop is finished and not to
        include its values in further entries of the experiment.

        This method is called by the loop itself if it ends its iterations,
        so is not typically needed by the user.
        """
        if loopHandler in self.loopsUnfinished:
            self.loopsUnfinished.remove(loopHandler)
    def _getAllParamNames(self):
        """Returns the attribute names of loop parameters (trialN etc)
        that the current set of loops contain, ready to build a wide-format
        data file.
        """
        names=copy.deepcopy(self._paramNamesSoFar)
        #get names (or identifiers) for all contained loops
        for thisLoop in self.loops:
            theseNames, vals = self._getLoopInfo(thisLoop)
            for name in theseNames:
                if name not in names:
                    names.append(name)
        return names
    def _getExtraInfo(self):
        """
        Get the names and vals from the extraInfo dict (if it exists)
        """
        if type(self.extraInfo) != dict:
            names=[]
            vals=[]
        else:
            names=self.extraInfo.keys()
            vals= self.extraInfo.values()
        return names, vals
    def _getLoopInfo(self, loop):
        """Returns the attribute names and values for the current trial of a particular loop.
        Does not return data inputs from the subject, only info relating to the trial
        execution.
        """
        names=[]
        vals=[]
        name = loop.name
        #standard attributes
        for attr in ['thisRepN', 'thisTrialN', 'thisN','thisIndex', 'stepSizeCurrent']:
            if hasattr(loop, attr):
                if attr=='stepSizeCurrent':
                    attrName=name+'.stepSize'
                else:
                    attrName = name+'.'+attr
                #append the attribute name and the current value
                names.append(attrName)
                vals.append(getattr(loop,attr))

        if hasattr(loop, 'thisTrial'):
            trial = loop.thisTrial
            if hasattr(trial,'items'):#is a TrialList object or a simple dict
                for attr,val in trial.items():
                    if attr not in self._paramNamesSoFar: self._paramNamesSoFar.append(attr)
                    names.append(attr)
                    vals.append(val)
            elif trial==[]:#we haven't had 1st trial yet? Not actually sure why this occasionally happens (JWP)
                pass
            else:
                names.append(name+'.thisTrial')
                vals.append(trial)
        elif hasattr(loop, 'intensities'):
            names.append(name+'.intensity')
            if len(loop.intensities)>0:
                vals.append(loop.intensities[-1])
            else:
                vals.append(None)

        return names, vals
    def addData(self, name, value):
        """Add the data with a given name to the current experiment.

        Typically the user does not need to use this function; if you added
        your data to the loop and had already added the loop to the
        experiment then the loop will automatically inform the experiment
        that it has received data.

        Multiple data name/value pairs can be added to any given entry of
        the data file and is considered part of the same entry until the
        nextEntry() call is made.

        e.g.::

            #add some data for this trial
            exp.addData('resp.rt', 0.8)
            exp.addData('resp.key', 'k')
            #end of trial - move to next line in data output
            exp.nextEntry()
        """
        if name not in self.dataNames:
            self.dataNames.append(name)
        self.thisEntry[name]=value

    def nextEntry(self):
        """Calling nextEntry indicates to the ExperimentHandler that the
        current trial has ended and so further
        addData() calls correspond to the next trial.
        """
        this=self.thisEntry
        #fetch data from each (potentially-nested) loop
        for thisLoop in self.loopsUnfinished:
            names, vals = self._getLoopInfo(thisLoop)
            for n, name in enumerate(names):
                this[name]=vals[n]
        #add the extraInfo dict to the data
        if type(self.extraInfo)==dict:
            this.update(self.extraInfo)#NB update() really means mergeFrom()
        self.entries.append(this)
        #then create new empty entry for n
        self.thisEntry = {}
    def saveAsWideText(self, fileName, delim=None,
                   matrixOnly=False,
                   appendFile=False):
        """Saves a long, wide-format text file, with one line representing the attributes and data
        for a single trial. Suitable for analysis in R and SPSS.

        If `appendFile=True` then the data will be added to the bottom of an existing file. Otherwise, if the file exists
        already it will be overwritten

        If `matrixOnly=True` then the file will not contain a header row, which can be handy if you want to append data
        to an existing file of the same format.
        """

        #create the file or print to stdout
        if appendFile: writeFormat='a'
        else: writeFormat='w' #will overwrite a file
        if os.path.exists(fileName) and writeFormat == 'w':
            logging.warning('Data file, %s, will be overwritten' %fileName)

        if fileName[-4:] in ['.csv', '.CSV']:
            delim=','
        else:
            delim='\t'

        if fileName=='stdout':
            f = sys.stdout
        elif fileName[-4:] in ['.csv', '.CSV','.dlm','.DLM', '.tsv','.TSV']:
            f= codecs.open(fileName,writeFormat, encoding = "utf-8")
        else:
            if delim==',':
                f= codecs.open(fileName+'.csv',writeFormat, encoding = "utf-8")
            else:
                f=codecs.open(fileName+'.dlm',writeFormat, encoding = "utf-8")

        names = self._getAllParamNames()
        names.extend(self.dataNames)
        names.extend(self._getExtraInfo()[0]) #names from the extraInfo dictionary
        #write a header line
        if not matrixOnly:
            for heading in names:
                f.write(u'%s%s' %(heading,delim))
            f.write('\n')
        #write the data for each entry
        for entry in self.entries:
            for name in names:
                if name in entry.keys():
                    if ',' in unicode(entry[name]):
                        f.write(u'"%s"%s' %(entry[name],delim))
                    else:
                        f.write(u'%s%s' %(entry[name],delim))
                else:
                    f.write(delim)
            f.write('\n')
        f.close()
        self.saveWideText=False
    def saveAsPickle(self,fileName, fileCollisionMethod = 'rename'):
        """Basically just saves a copy of self (with data) to a pickle file.

        This can be reloaded if necessary and further analyses carried out.

        :Parameters:

            fileCollisionMethod: Collision method passed to ~psychopy.misc._handleFileCollision
        """
        #otherwise use default location
        if not fileName.endswith('.psydat'):
            fileName+='.psydat'
        if os.path.exists(fileName):
            fileName = misc._handleFileCollision(fileName, fileCollisionMethod)

        #create the file or print to stdout
        f = open(fileName, 'wb')
        cPickle.dump(self, f)
        f.close()
        #no need to save again
        self.savePickle=False

    def abort(self):
        """Inform the ExperimentHandler that the run was aborted.

        Experiment handler will attempt automatically to save data (even in the event of a crash if possible).
        So if you quit your script early you may want to tell the Handler not to save out the data files for this run.
        This is the method that allows you to do that.
        """
        self.savePickle=False
        self.saveWideText=False

class TrialType(dict):
    """This is just like a dict, except that you can access keys with obj.key
    """
    def __getattribute__(self, name):
        try:#to get attr from dict in normal way (passing self)
            return dict.__getattribute__(self, name)
        except AttributeError:
            try:
                return self[name]
            except KeyError:
                raise AttributeError, ('TrialType has no attribute (or key) \'%s\'' %(name))

class _BaseTrialHandler(object):
    def setExp(self, exp):
        """Sets the ExperimentHandler that this handler is attached to

        Do NOT attempt to set the experiment using::

            trials._exp = myExperiment

        because it needs to be performed using the `weakref` module.
        """
        #need to use a weakref to avoid creating a circular reference that
        #prevents effective object deletion
        expId=id(exp)
        _experiments[expId] = exp
        self._exp = expId
    def getExp(self):
        """Return the ExperimentHandler that this handler is attached to, if any.
        Returns None if not attached
        """
        if self._exp==None or self._exp not in _experiments:
            return None
        else:
            return _experiments[self._exp]
    def _terminate(self):
        """Remove references to ourself in experiments and terminate the loop
        """
        #remove ourself from the list of unfinished loops in the experiment
        exp=self.getExp()
        if exp!=None:
            exp.loopEnded(self)
        #and halt the loop
        raise StopIteration
    def saveAsPickle(self,fileName, fileCollisionMethod = 'rename'):
        """Basically just saves a copy of the handler (with data) to a pickle file.

        This can be reloaded if necessesary and further analyses carried out.

        :Parameters:

            fileCollisionMethod: Collision method passed to ~psychopy.misc._handleFileCollision
        """
        if self.thisTrialN<1 and self.thisRepN<1:#if both are <1 we haven't started
            logging.info('.saveAsPickle() called but no trials completed. Nothing saved')
            return -1
        #otherwise use default location
        if not fileName.endswith('.psydat'):
            fileName+='.psydat'
        if os.path.exists(fileName):
            fileName = misc._handleFileCollision(fileName, fileCollisionMethod)

        #create the file or print to stdout
        f = open(fileName, 'wb')
        cPickle.dump(self, f)
        f.close()
    def saveAsText(self,fileName,
                   stimOut=[],
                   dataOut=('n','all_mean','all_std', 'all_raw'),
                   delim=None,
                   matrixOnly=False,
                   appendFile=True,
                   summarised=True,
                   ):
        """
        Write a text file with the data and various chosen stimulus attributes

         :Parameters:

            fileName:
                will have .dlm appended (so you can double-click it to
                open in excel) and can include path info.

            stimOut:
                the stimulus attributes to be output. To use this you need to
                use a list of dictionaries and give here the names of dictionary keys
                that you want as strings

            dataOut:
                a list of strings specifying the dataType and the analysis to
                be performed,in the form `dataType_analysis`. The data can be any of the types that
                you added using trialHandler.data.add() and the analysis can be either
                'raw' or most things in the numpy library, including;
                'mean','std','median','max','min'...
                The default values will output the raw, mean and std of all datatypes found

            delim:
                allows the user to use a delimiter other than tab ("," is popular with file extension ".csv")

            matrixOnly:
                outputs the data with no header row or extraInfo attached

            appendFile:
                will add this output to the end of the specified file if it already exists

        """
        if self.thisTrialN<1 and self.thisRepN<1:#if both are <1 we haven't started
            logging.info('TrialHandler.saveAsText called but no trials completed. Nothing saved')
            return -1

        dataArray = self._createOutputArray(stimOut=[],
            dataOut=dataOut,
            matrixOnly=matrixOnly)

        #set default delimiter if none given
        if delim==None:
            if fileName[-4:] in ['.csv','.CSV']:
                delim=','
            else:
                delim='\t'

        #create the file or print to stdout
        if appendFile: writeFormat='a'
        else: writeFormat='w' #will overwrite a file
        if fileName=='stdout':
            f = sys.stdout
        elif fileName[-4:] in ['.dlm','.DLM', '.csv', '.CSV']:
            f= codecs.open(fileName,writeFormat, encoding = "utf-8")
        else:
            if delim==',':
                f= codecs.open(fileName+'.csv',writeFormat, encoding = "utf-8")
            else:
                f=codecs.open(fileName+'.dlm',writeFormat, encoding = "utf-8")

        #loop through lines in the data matrix
        for line in dataArray:
            for cellN, entry in enumerate(line):
                if delim in unicode(entry):#surround in quotes to prevent effect of delimiter
                    f.write(u'"%s"' %unicode(entry))
                else:
                    f.write(unicode(entry))
                if cellN<(len(line)-1):
                    f.write(delim)
            f.write("\n")#add an EOL at end of each line
        if f != sys.stdout:
            f.close()
            logging.info('saved data to %s' %f.name)
    def printAsText(self, stimOut=[],
                    dataOut=('all_mean', 'all_std', 'all_raw'),
                    delim='\t',
                    matrixOnly=False,
                  ):
        """Exactly like saveAsText() except that the output goes
        to the screen instead of a file"""
        self.saveAsText('stdout', stimOut, dataOut, delim, matrixOnly)

    def saveAsExcel(self,fileName, sheetName='rawData',
                    stimOut=[],
                    dataOut=('n','all_mean','all_std', 'all_raw'),
                    matrixOnly=False,
                    appendFile=True,
                    ):
        """
        Save a summary data file in Excel OpenXML format workbook (:term:`xlsx`) for processing
        in most spreadsheet packages. This format is compatible with
        versions of Excel (2007 or greater) and and with OpenOffice (>=3.0).

        It has the advantage over the simpler text files (see :func:`TrialHandler.saveAsText()` )
        that data can be stored in multiple named sheets within the file. So you could have a single file
        named after your experiment and then have one worksheet for each participant. Or you could have
        one file for each participant and then multiple sheets for repeated sessions etc.

        The file extension `.xlsx` will be added if not given already.

        :Parameters:

            fileName: string
                the name of the file to create or append. Can include relative or absolute path

            sheetName: string
                the name of the worksheet within the file

            stimOut: list of strings
                the attributes of the trial characteristics to be output. To use this you need to have provided
                a list of dictionaries specifying to trialList parameter of the TrialHandler
                and give here the names of strings specifying entries in that dictionary

            dataOut: list of strings
                specifying the dataType and the analysis to
                be performed, in the form `dataType_analysis`. The data can be any of the types that
                you added using trialHandler.data.add() and the analysis can be either
                'raw' or most things in the numpy library, including
                'mean','std','median','max','min'. e.g. `rt_max` will give a column of max reaction
                times across the trials assuming that `rt` values have been stored.
                The default values will output the raw, mean and std of all datatypes found

            appendFile: True or False
                If False any existing file with this name will be overwritten. If True then a new worksheet will be appended.
                If a worksheet already exists with that name a number will be added to make it unique.


        """

        if self.thisTrialN<1 and self.thisRepN<1:#if both are <1 we haven't started
            logging.info('TrialHandler.saveAsExcel called but no trials completed. Nothing saved')
            return -1

        #NB this was based on the limited documentation (1 page wiki) for openpyxl v1.0
        if not haveOpenpyxl:
            raise ImportError, 'openpyxl is required for saving files in Excel (xlsx) format, but was not found.'
            return -1

        #create the data array to be sent to the Excel file
        dataArray = self._createOutputArray(stimOut=[],
            dataOut=dataOut,
            matrixOnly=matrixOnly)

        #import necessary subpackages - they are small so won't matter to do it here
        from openpyxl.workbook import Workbook
        from openpyxl.writer.excel import ExcelWriter
        from openpyxl.reader.excel import load_workbook

        if not fileName.endswith('.xlsx'): fileName+='.xlsx'
        #create or load the file
        if appendFile and os.path.isfile(fileName):
            wb = load_workbook(fileName)
            newWorkbook=False
        else:
            if not appendFile: #the file exists but we're not appending, so will be overwritten
                logging.warning('Data file, %s, will be overwritten' %fileName)
            wb = Workbook()#create new workbook
            wb.properties.creator='PsychoPy'+psychopy.__version__
            newWorkbook=True

        ew = ExcelWriter(workbook = wb)

        if newWorkbook:
            ws = wb.worksheets[0]
            ws.title=sheetName
        else:
            ws=wb.create_sheet()
            ws.title=sheetName

        #loop through lines in the data matrix
        for lineN, line in enumerate(dataArray):
            if line==None:
                continue
            for colN, entry in enumerate(line):
                if entry in [None]:
                    entry=''
                try:
                    ws.cell(_getExcelCellName(col=colN,row=lineN)).value = float(entry)#if it can conver to a number (from numpy) then do it
                except:
                    ws.cell(_getExcelCellName(col=colN,row=lineN)).value = unicode(entry)#else treat as unicode

        ew.save(filename = fileName)

    def nextTrial(self):
        """DEPRECATION WARNING: nextTrial() will be deprecated
        please use next() instead.
        jwp: 19/6/06
        """
        if self._warnUseOfNext:
            logging.warning("""DEPRECATION WARNING: nextTrial() will be deprecated
        please use next() instead.
        jwp: 19/6/06
        """)
            self._warnUseOfNext=False
        return self.next()
    def getOriginPathAndFile(self, originPath=None):
        """Attempts to determine the path of the script that created this data file
        and returns both the path to that script and it's contents.
        Useful to store the entire experiment with the data.

        If originPath is provided (e.g. from Builder) then this is used otherwise
        the calling script is the originPath (fine from a standard python script).
        """
        #self.originPath and self.origin (the contents of the origin file)
        if originPath==None or not os.path.isfile(originPath):
            try:
                originPath = inspect.getouterframes(inspect.currentframe())[1][1]
                logging.debug("Using %s as origin file" %originPath)
            except:
                logging.debug("Failed to find origin file using inspect.getouterframes")
                return '',''
        if os.path.isfile(originPath):#do we NOW have a path?
            origin = codecs.open(originPath,"r", encoding = "utf-8").read()
        else:
            origin=None
        return originPath, origin

class TrialHandler(_BaseTrialHandler):
    """Class to handle trial sequencing and data storage.

    Calls to .next() will fetch the next trial object given to this handler,
    according to the method specified (random, sequential, fullRandom). Calls
    will raise a StopIteration error if trials have finished.

    See demo_trialHandler.py

    The psydat file format is literally just a pickled copy of the TrialHandler object that
    saved it. You can open it with::

            from psychopy import misc
            dat = misc.fromFile(path)

    Then you'll find that `dat` has the following attributes that
    """
    def __init__(self,
                 trialList,
                 nReps,
                 method='random',
                 dataTypes=None,
                 extraInfo=None,
                 seed=None,
                 originPath=None,
                 name=''):
        """

        :Parameters:

            trialList: a simple list (or flat array) of dictionaries specifying conditions
                This can be imported from an excel/csv file using :func:`~psychopy.data.importConditions`

            nReps: number of repeats for all conditions

            method: *'random',* 'sequential', or 'fullRandom'
                'sequential' obviously presents the conditions in the order they appear in the list.
                'random' will result in a shuffle of the conditions on each repeat, but all conditions
                occur once before the second repeat etc. 'fullRandom' fully randomises the
                trials across repeats as well, which means you could potentially run all trials of
                one condition before any trial of another.

            dataTypes: (optional) list of names for data storage. e.g. ['corr','rt','resp']
                If not provided then these will be created as needed during calls to
                :func:`~psychopy.data.TrialHandler.addData`

            extraInfo: A dictionary
                This will be stored alongside the data and usually describes the experiment and
                subject ID, date etc.

            seed: an integer
                If provided then this fixes the random number generator to use the same pattern
                of trials, by seeding its startpoint

            originPath: a string describing the location of the script/experiment file path
                The psydat file format will store a copy of the experiment if possible. If no file path
                is provided here then the TrialHandler will still store a copy of the script where it was
                created

        :Attributes (after creation):

            .data - a dictionary of numpy arrays, one for each data type stored

            .trialList - the original list of dicts, specifying the conditions

            .thisIndex - the index of the current trial in the original conditions list

            .nTotal - the total number of trials that will be run

            .nRemaining - the total number of trials remaining

            .thisN - total trials completed so far

            .thisRepN - which repeat you are currently on

            .thisTrialN - which trial number *within* that repeat

            .thisTrial - a dictionary giving the parameters of the current trial

            .finished - True/False for have we finished yet

            .extraInfo - the dictionary of extra info as given at beginning

            .origin - the contents of the script or builder experiment that created the handler

        """
        self.name=name
        if trialList in [None, []]:#user wants an empty trialList
            self.trialList = [None]#which corresponds to a list with a single empty entry
        else:
            self.trialList =trialList
        #convert any entry in the TrialList into a TrialType object (with obj.key or obj[key] access)
        for n, entry in enumerate(trialList):
            if type(entry)==dict:
                trialList[n]=TrialType(entry)
        self.nReps = int(nReps)
        self.nTotal = self.nReps*len(self.trialList)
        self.nRemaining =self.nTotal #subtract 1 each trial
        self.method = method
        self.thisRepN = 0        #records which repetition or pass we are on
        self.thisTrialN = -1    #records which trial number within this repetition
        self.thisN = -1
        self.thisIndex = 0        #the index of the current trial in the conditions list
        self.thisTrial = []
        self.finished=False
        self.extraInfo=extraInfo
        self._warnUseOfNext=True
        self.seed=seed
        #create dataHandler
        self.data = DataHandler(trials=self)
        if dataTypes!=None:
            self.data.addDataType(dataTypes)
        self.data.addDataType('ran')
        self.data['ran'].mask=False#this is a bool - all entries are valid
        self.data.addDataType('order')
        #generate stimulus sequence
        if self.method in ['random','sequential', 'fullRandom']:
            self.sequenceIndices = self._createSequence()
        else: self.sequenceIndices=[]

        self.originPath, self.origin = self.getOriginPathAndFile(originPath)
        self._exp = None#the experiment handler that owns me!

    def __iter__(self):
        return self
    def __repr__(self):
        """prints a more verbose version of self as string"""
        return self.__str__(verbose=True)

    def __str__(self, verbose=False):
        """string representation of the object"""
        strRepres = 'psychopy.data.TrialHandler(\n'
        attribs = dir(self)

        #print data first, then all others
        try: data=self.data
        except: data=None
        if data:
            strRepres += str('\tdata=')
            strRepres +=str(data)+'\n'

        for thisAttrib in attribs:
            #can handle each attribute differently
            if 'instancemethod' in str(type(getattr(self,thisAttrib))):
                #this is a method
                continue
            elif thisAttrib[0]=='_':
                #the attrib is private
                continue
            elif thisAttrib=='data':
                #we handled this first
                continue
            elif len(str(getattr(self,thisAttrib)))>20 and \
                 not verbose:
                #just give type of LONG public attribute
                strRepres += str('\t'+thisAttrib+'=')
                strRepres += str(type(getattr(self,thisAttrib)))+'\n'
            else:
                #give the complete contents of attribute
                strRepres += str('\t'+thisAttrib+'=')
                strRepres += str(getattr(self,thisAttrib))+'\n'

        strRepres+=')'
        return strRepres

    def _createSequence(self):
        """
        Pre-generates the sequence of trial presentations (for non-adaptive methods).
        This is called automatically when the TrialHandler is initialised so doesn't
        need an explicit call from the user.

        The returned sequence has form indices[stimN][repN]
        Example: sequential with 6 trialtypes (rows), 5 reps (cols), returns:
            [[0 0 0 0 0]
             [1 1 1 1 1]
             [2 2 2 2 2]
             [3 3 3 3 3]
             [4 4 4 4 4]
             [5 5 5 5 5]]
        These 30 trials will be returned by .next() in the order:
            0, 1, 2, 3, 4, 5,   0, 1, 2, ...  ... 3, 4, 5

        To add a new type of sequence (as of v1.65.02):
        - add the sequence generation code here
        - adjust "if self.method in [ ...]:" in both __init__ and .next()
        - adjust allowedVals in experiment.py -> shows up in DlgLoopProperties
        Note that users can make any sequence whatsoever outside of PsychoPy, and
        specify sequential order; any order is possible this way.
        """
        # create indices for a single rep
        indices = numpy.asarray(self._makeIndices(self.trialList), dtype=int)

        if self.method == 'random':
            sequenceIndices = []
            seed=self.seed
            for thisRep in range(self.nReps):
                thisRepSeq = misc.shuffleArray(indices.flat, seed=seed).tolist()
                seed=None#so that we only seed the first pass through!
                sequenceIndices.append(thisRepSeq)
            sequenceIndices = numpy.transpose(sequenceIndices)
        elif self.method == 'sequential':
            sequenceIndices = numpy.repeat(indices,self.nReps,1)
        elif self.method == 'fullRandom':
            # indices*nReps, flatten, shuffle, unflatten; only use seed once
            sequential = numpy.repeat(indices, self.nReps,1) # = sequential
            randomFlat = misc.shuffleArray(sequential.flat, seed=self.seed)
            sequenceIndices = numpy.reshape(randomFlat, (len(indices), self.nReps))
        logging.exp('Created sequence: %s, trialTypes=%d, nReps=%i, seed=%s' %
                (self.method, len(indices), self.nReps, str(self.seed) )  )
        return sequenceIndices

    def _makeIndices(self,inputArray):
        """
        Creates an array of tuples the same shape as the input array
        where each tuple contains the indices to itself in the array.

        Useful for shuffling and then using as a reference.
        """
        inputArray  = numpy.asarray(inputArray, 'O')#make sure its an array of objects (can be strings etc)
        #get some simple variables for later
        dims=inputArray.shape
        dimsProd=numpy.product(dims)
        dimsN = len(dims)
        dimsList = range(dimsN)
        listOfLists = []
        arrayOfTuples = numpy.ones(dimsProd, 'O')#this creates space for an array of any objects

        #for each dimension create list of its indices (using modulo)
        for thisDim in dimsList:
            prevDimsProd = numpy.product(dims[:thisDim])
            thisDimVals = numpy.arange(dimsProd)/prevDimsProd % dims[thisDim] #NB this means modulus in python
            listOfLists.append(thisDimVals)

        #convert to array
        indexArr = numpy.asarray(listOfLists)
        for n in range(dimsProd):
            arrayOfTuples[n] = tuple((indexArr[:,n]))
        return (numpy.reshape(arrayOfTuples,dims)).tolist()

    def next(self):
        """Advances to next trial and returns it.
        Updates attributes; thisTrial, thisTrialN and thisIndex
        If the trials have ended this method will raise a StopIteration error.
        This can be handled with code such as::

            trials = TrialHandler(.......)
            for eachTrial in trials:#automatically stops when done
                #do stuff

        or::

            trials = TrialHandler(.......)
            while True: #ie forever
                try:
                    thisTrial = trials.next()
                except StopIteration:#we got a StopIteration error
                    break #break out of the forever loop
                #do stuff here for the trial
        """
        #update pointer for next trials
        self.thisTrialN+=1#number of trial this pass
        self.thisN+=1 #number of trial in total
        self.nRemaining-=1
        if self.thisTrialN==len(self.trialList):
            #start a new repetition
            self.thisTrialN=0
            self.thisRepN+=1
        if self.thisRepN>=self.nReps:
            #all reps complete
            self.thisTrial=[]
            self.finished=True

        if self.finished==True:
            self._terminate()

        #fetch the trial info
        if self.method in ['random','sequential','fullRandom']:
            self.thisIndex = self.sequenceIndices[self.thisTrialN][self.thisRepN]
            self.thisTrial = self.trialList[self.thisIndex]
            self.data.add('ran',1)
            self.data.add('order',self.thisN)
        logging.exp('New trial (rep=%i, index=%i): %s' %(self.thisRepN, self.thisTrialN, self.thisTrial), obj=self.thisTrial)
        return self.thisTrial

    def getFutureTrial(self, n=1):
        """Returns the condition for n trials into the future without advancing
        the trials.
        """
        if n>self.nRemaining:
            return None
        seqs = numpy.array(self.sequenceIndices).transpose().flat
        condIndex=seqs[self.thisN+n]
        return self.trialList[condIndex]

    def _createOutputArray(self,stimOut,dataOut,delim=None,
                          matrixOnly=False):
        """
        Does the leg-work for saveAsText and saveAsExcel.
        Combines stimOut with ._parseDataOutput()
        """
        if stimOut==[] and len(self.trialList) and hasattr(self.trialList[0],'keys'):
            stimOut=self.trialList[0].keys()
            #these get added somewhere (by DataHandler?)
            if 'n' in stimOut:
                stimOut.remove('n')
            if 'float' in stimOut:
                stimOut.remove('float')

        lines=[]
        #parse the dataout section of the output
        dataOut, dataAnal, dataHead = self._createOutputArrayData(dataOut=dataOut)
        if not matrixOnly:
            thisLine=[]
            lines.append(thisLine)
            #write a header line
            for heading in stimOut+dataHead:
                if heading=='ran_sum': heading ='n'
                elif heading=='order_raw': heading ='order'
                thisLine.append(heading)

        #loop through stimuli, writing data
        for stimN in range(len(self.trialList)):
            thisLine=[]
            lines.append(thisLine)
            #first the params for this stim (from self.trialList)
            for heading in stimOut:
                thisLine.append(self.trialList[stimN][heading])

            #then the data for this stim (from self.data)
            for thisDataOut in dataOut:
                #make a string version of the data and then format it
                tmpData = dataAnal[thisDataOut][stimN]
                if hasattr(tmpData,'tolist'): #is a numpy array
                    strVersion = unicode(tmpData.tolist())
                    #for numeric data replace None with a blank cell
                    if tmpData.dtype.kind not in ['SaUV']:
                        strVersion=strVersion.replace('None','')
                elif tmpData in [None,'None']:
                    strVersion=''
                else:
                    strVersion = unicode(tmpData)

                if strVersion=='()':
                    strVersion="--"# 'no data' in masked array should show as "--"
                #handle list of values (e.g. rt_raw )
                if len(strVersion) and strVersion[0] in ["[", "("] and strVersion[-1] in ["]", ")"]:
                    strVersion=strVersion[1:-1]#skip first and last chars
                #handle lists of lists (e.g. raw of multiple key presses)
                if len(strVersion) and strVersion[0] in ["[", "("] and strVersion[-1] in ["]", ")"]:
                    tup = eval(strVersion) #convert back to a tuple
                    for entry in tup:
                        #contents of each entry is a list or tuple so keep in quotes to avoid probs with delim
                        thisLine.append(unicode(entry))
                else:
                    thisLine.extend(strVersion.split(','))

        #add self.extraInfo
        if (self.extraInfo != None) and not matrixOnly:
            lines.append([])
            lines.append(['extraInfo'])#give a single line of space and then a heading
            for key, value in self.extraInfo.items():
                lines.append([key,value])
        return lines

    def _createOutputArrayData(self, dataOut):
        """This just creates the dataOut part of the output matrix.
        It is called by _createOutputArray() which creates the header line and adds the stimOut columns
        """
        dataHead=[]#will store list of data headers
        dataAnal=dict([])    #will store data that has been analyzed
        if type(dataOut)==str: dataOut=[dataOut]#don't do list convert or we get a list of letters
        elif type(dataOut)!=list: dataOut = list(dataOut)

        #expand any 'all' dataTypes to be the full list of available dataTypes
        allDataTypes=self.data.keys()
        #treat these separately later
        allDataTypes.remove('ran')
        #ready to go trhough standard data types
        dataOutNew=[]
        for thisDataOut in dataOut:
            if thisDataOut=='n':
                #n is really just the sum of the ran trials
                dataOutNew.append('ran_sum')
                continue#no need to do more with this one
            #then break into dataType and analysis
            dataType, analType =string.rsplit(thisDataOut, '_', 1)
            if dataType=='all':
                dataOutNew.extend([key+"_"+analType for key in allDataTypes])
                if 'order_mean' in dataOutNew: dataOutNew.remove('order_mean')
                if 'order_std' in dataOutNew: dataOutNew.remove('order_std')
            else:
                dataOutNew.append(thisDataOut)
        dataOut=dataOutNew
        dataOut.sort()#so that all datatypes come together, rather than all analtypes

        #do the various analyses, keeping track of fails (e.g. mean of a string)
        dataOutInvalid=[]
        #add back special data types (n and order)
        if 'ran_sum' in dataOut:#move n to the first column
            dataOut.remove('ran_sum')
            dataOut.insert(0,'ran_sum')
        if 'order_raw' in dataOut:#move order_raw to the second column
            dataOut.remove('order_raw')
            dataOut.append('order_raw')
        #do the necessary analysis on the data
        for thisDataOutN,thisDataOut in enumerate(dataOut):
            dataType, analType =string.rsplit(thisDataOut, '_', 1)
            if not self.data.has_key(dataType):
                dataOutInvalid.append(thisDataOut)#that analysis can't be done
                continue
            thisData = self.data[dataType]

            #set the header
            dataHead.append(dataType+'_'+analType)
            #analyse thisData using numpy module
            if analType in dir(numpy):
                try:#this will fail if we try to take mean of a string for example
                    if analType=='std':
                        thisAnal = numpy.std(thisData,axis=1,ddof=0)
                        #normalise by N-1 instead. his should work by setting ddof=1
                        #but doesn't as of 08/2010 (because of using a masked array?)
                        N=thisData.shape[1]
                        if N == 1:
                            thisAnal*=0 #prevent a divide-by-zero error
                        else:
                            thisAnal = thisAnal*numpy.sqrt(N)/numpy.sqrt(N-1)
                    else:
                        exec("thisAnal = numpy.%s(thisData,1)" %analType)
                except:
                    dataHead.remove(dataType+'_'+analType)#that analysis doesn't work
                    dataOutInvalid.append(thisDataOut)
                    continue#to next analysis
            elif analType=='raw':
                thisAnal=thisData
            else:
                raise AttributeError, 'You can only use analyses from numpy'
            #add extra cols to header if necess
            if len(thisAnal.shape)>1:
                for n in range(thisAnal.shape[1]-1):
                    dataHead.append("")
            dataAnal[thisDataOut]=thisAnal

        #remove invalid analyses (e.g. average of a string)
        for invalidAnal in dataOutInvalid: dataOut.remove(invalidAnal)
        return dataOut, dataAnal, dataHead


    def saveAsWideText(self,fileName,
                   delim='\t',
                   matrixOnly=False,
                   appendFile=True,
                  ):
        """
        Write a text file with the session, stimulus, and data values from each trial in chronological order.

        That is, unlike 'saveAsText' and 'saveAsExcel':
         - each row comprises information from only a single trial.
         - no summarising is done (such as collapsing to produce mean and standard deviation values across trials).

        This 'wide' format, as expected by R for creating dataframes, and various other analysis programs, means that some
        information must be repeated on every row.

        In particular, if the trialHandler's 'extraInfo' exists, then each entry in there occurs in every row.
        In builder, this will include any entries in the 'Experiment info' field of the 'Experiment settings' dialog.
        In Coder, this information can be set using something like:
            myTrialHandler.extraInfo = {'SubjID':'Joan Smith', 'DOB':1970 Nov 16, 'Group':'Control'}

         :Parameters:

            fileName:
                if extension is not specified, '.csv' will be appended if the delimiter is ',', else '.txt' will be appended.
                Can include path info.

            delim:
                allows the user to use a delimiter other than the default tab ("," is popular with file extension ".csv")

            matrixOnly:
                outputs the data with no header row.

            appendFile:
                will add this output to the end of the specified file if it already exists.

        """
        if self.thisTrialN<1 and self.thisRepN<1:#if both are <1 we haven't started
            logging.info('TrialHandler.saveAsWideText called but no trials completed. Nothing saved')
            return -1

        #create the file or print to stdout
        if appendFile:
            writeFormat='a'
        else: writeFormat='w' #will overwrite a file
        if fileName=='stdout':
            f = sys.stdout
        elif fileName[-4:] in ['.dlm','.DLM', '.tsv', '.TSV', '.txt', '.TXT', '.csv', '.CSV']:
            f = codecs.open(fileName,writeFormat, encoding = "utf-8")
        else:
            if delim==',': f = codecs.open(fileName+'.csv', writeFormat, encoding="utf-8")
            else: f=codecs.open(fileName+'.txt',writeFormat, encoding = "utf-8")

        # collect parameter names related to the stimuli:
        if self.trialList[0]:
            header = self.trialList[0].keys()
        else:
            header = []
        # and then add parameter names related to data (e.g. RT)
        header.extend(self.data.dataTypes)

        # loop through each trial, gathering the actual values:
        dataOut = []
        trialCount = 0
        # total number of trials = number of trialtypes * number of repetitions:

        repsPerType={}
        for rep in range(self.nReps):
            for trialN in range(len(self.trialList)):
                #find out what trial type was on this trial
                trialTypeIndex = self.sequenceIndices[trialN, rep]
                #determine which repeat it is for this trial
                if trialTypeIndex not in repsPerType.keys():
                    repsPerType[trialTypeIndex]=0
                else:
                    repsPerType[trialTypeIndex]+=1
                repThisType=repsPerType[trialTypeIndex]#what repeat are we on for this trial type?

                # create a dictionary representing each trial:
                # this is wide format, so we want fixed information (e.g. subject ID, date, etc) repeated every line if it exists:
                if (self.extraInfo != None):
                    nextEntry = self.extraInfo.copy()
                else:
                    nextEntry = {}

                # add a trial number so the original order of the data can always be recovered if sorted during analysis:
                trialCount += 1
                nextEntry["TrialNumber"] = trialCount

                # now collect the value from each trial of the variables named in the header:
                for parameterName in header:
                    # the header includes both trial and data variables, so need to check before accessing:
                    if self.trialList[trialTypeIndex] and self.trialList[trialTypeIndex].has_key(parameterName):
                        nextEntry[parameterName] = self.trialList[trialTypeIndex][parameterName]
                    elif self.data.has_key(parameterName):
                        nextEntry[parameterName] = self.data[parameterName][trialTypeIndex][repThisType]
                    else: # allow a null value if this parameter wasn't explicitly stored on this trial:
                        nextEntry[parameterName] = ''

                #store this trial's data
                dataOut.append(nextEntry)

        # get the extra 'wide' parameter names into the header line:
        header.insert(0,"TrialNumber")
        if (self.extraInfo != None):
            for key in self.extraInfo:
                header.insert(0, key)

        if not matrixOnly:
        # write the header row:
            nextLine = ''
            for parameterName in header:
                nextLine = nextLine + parameterName + delim
            f.write(nextLine[:-1] + '\n') # remove the final orphaned tab character

        # write the data matrix:
        for trial in dataOut:
            nextLine = ''
            for parameterName in header:
                nextLine = nextLine + unicode(trial[parameterName]) + delim
            nextLine = nextLine[:-1] # remove the final orphaned tab character
            f.write(nextLine + '\n')

        if f != sys.stdout:
            f.close()
            logging.info('saved wide-format data to %s' %f.name)

    def addData(self, thisType, value, position=None):
        """Add data for the current trial
        """
        self.data.add(thisType, value, position=None)
        if self.getExp()!=None:#update the experiment handler too
            self.getExp().addData(thisType, value)


def importTrialTypes(fileName, returnFieldNames=False):
    """importTrialTypes is DEPRECATED (as of v1.70.00)
    Please use `importConditions` for identical functionality.
    """
    logging.warning("importTrialTypes is DEPRECATED (as of v1.70.00). Please use `importConditions` for identical functionality.")
    return importConditions(fileName, returnFieldNames)

def importConditions(fileName, returnFieldNames=False):
    """Imports a list of conditions from an .xlsx, .csv, or .pkl file

    The output is suitable as an input to :class:`TrialHandler` `trialTypes` or to
    :class:`MultiStairHandler` as a `conditions` list.

    If `fileName` ends with:
        - .csv:  import as a comma-separated-value file (header + row x col)
        - .xlsx: import as Excel 2007 (xlsx) files. Sorry no support for older (.xls) is planned.
        - .pkl:  import from a pickle file as list of lists (header + row x col)

    The file should contain one row per type of trial needed and one column
    for each parameter that defines the trial type. The first row should give
    parameter names, which should:

        - be unique
        - begin with a letter (upper or lower case)
        - contain no spaces or other punctuation (underscores are permitted)

    """
    def _assertValidVarNames(fieldNames, fileName):
        """screens a list of names as candidate variable names. if all names are
        OK, return silently; else raise ImportError with msg
        """
        if not all(fieldNames):
            raise ImportError, 'Conditions file %s: Missing parameter name(s); empty cell(s) in the first row?' % fileName
        for name in fieldNames:
            OK, msg = isValidVariableName(name)
            if not OK: #tailor message to importConditions
                msg = msg.replace('Variables', 'Parameters (column headers)')
                raise ImportError, 'Conditions file %s: %s%s"%s"' %(fileName, msg, os.linesep*2, name)

    if fileName in ['None','none',None]:
        if returnFieldNames:
            return [], []
        return []
    if not os.path.isfile(fileName):
        raise ImportError, 'Conditions file not found: %s' %os.path.abspath(fileName)

    if fileName.endswith('.csv'):
        #use csv import library to fetch the fieldNames
        f = open(fileName, 'rU')#the U converts line endings to os.linesep (not unicode!)
        #lines = f.read().split(os.linesep)#csv module is temperamental with line endings
        try:
            reader = csv.reader(f)#.split(os.linesep))
        except:
            raise ImportError, 'Could not open %s as conditions' % fileName
        fieldNames = reader.next() # first row
        _assertValidVarNames(fieldNames, fileName)
        #use matplotlib to import data and intelligently check for data types
        #all data in one column will be given a single type (e.g. if one cell is string, all will be set to string)
        trialsArr = mlab.csv2rec(f) # data = non-header row x col
        f.close()
        #convert the record array into a list of dicts
        trialList = []
        for trialN, trialType in enumerate(trialsArr):
            thisTrial ={}
            for fieldN, fieldName in enumerate(fieldNames):
                val = trialsArr[trialN][fieldN]
                if type(val)==numpy.string_:
                    val = unicode(val.decode('utf-8'))
                    #if it looks like a list, convert it:
                    if val.startswith('[') and val.endswith(']'):
                        #exec('val=%s' %unicode(val.decode('utf8')))
                        val = eval(val)
                thisTrial[fieldName] = val
            trialList.append(thisTrial)
    elif fileName.endswith('.pkl'):
        f = open(fileName, 'rU') # is U needed?
        try:
            trialsArr = cPickle.load(f)
        except:
            raise ImportError, 'Could not open %s as conditions' % fileName
        f.close()
        trialList = []
        fieldNames = trialsArr[0] # header line first
        _assertValidVarNames(fieldNames, fileName)
        for row in trialsArr[1:]:
            thisTrial = {}
            for fieldN, fieldName in enumerate(fieldNames):
                thisTrial[fieldName] = row[fieldN] # type is correct, being .pkl
            trialList.append(thisTrial)
    else:
        if not haveOpenpyxl:
            raise ImportError, 'openpyxl is required for loading excel format files, but it was not found.'
        try:
            wb = load_workbook(filename = fileName)
        except: # InvalidFileException(unicode(e)): # this fails
            raise ImportError, 'Could not open %s as conditions' % fileName
        ws = wb.worksheets[0]
        nCols = ws.get_highest_column()
        nRows = ws.get_highest_row()

        #get parameter names from the first row header
        fieldNames = []
        for colN in range(nCols):
            fieldName = ws.cell(_getExcelCellName(col=colN, row=0)).value
            fieldNames.append(fieldName)
        _assertValidVarNames(fieldNames, fileName)

        #loop trialTypes
        trialList = []
        for rowN in range(1, nRows):#skip header first row
            thisTrial={}
            for colN in range(nCols):
                val = ws.cell(_getExcelCellName(col=colN, row=rowN)).value
                #if it looks like a list, convert it
                if type(val) in [unicode, str] and (
                        val.startswith('[') and val.endswith(']') or
                        val.startswith('(') and val.endswith(')') ):
                    val = eval(val)
                fieldName = fieldNames[colN]
                thisTrial[fieldName] = val
            trialList.append(thisTrial)

    logging.exp('Imported %s as conditions, %d conditions, %d params' %
                 (fileName, len(trialList), len(fieldNames)))
    if returnFieldNames:
        return (trialList,fieldNames)
    else:
        return trialList

def createFactorialTrialList(factors):
    """Create a trialList by entering a list of factors with names (keys) and levels (values)
    it will return a trialList in which all factors have been factorially combined (so for example
    if there are two factors with 3 and 5 levels the trialList will be a list of 3*5 = 15, each specifying
    the values for a given trial

    Usage::

        trialList = createFactorialTrialList(factors)

    :Parameters:

        factors : a dictionary with names (keys) and levels (values) of the factors

    Example::

        mytrials = createFactorialTrialList( factors={"text": ["red", "green", "blue"],
            "letterColor": ["red", "green"], "size": [0,1]})
    """

    # the first step is to place all the factorial combinations in a list of lists
    tempListOfLists=[[]]
    for key in factors:
        alist = factors[key]   # this takes the levels of each factor as a set of values (a list) at a time
        tempList = []
        for value in alist:     # now we loop over the values in a given list, and add each value of the other lists
            for iterList in tempListOfLists:
                tempList.append(iterList + [key,value])
        tempListOfLists = tempList

    # this second step is so we can return a list in the format of trialList
    trialList = []
    for atrial in tempListOfLists:
        keys = atrial[0::2]          #the even elements are keys
        values = atrial[1::2]       #the odd elements are values
        atrialDict = {}
        for i in range(len(keys)):
            atrialDict[keys[i]] = values[i]     #this combines the key with the value
        trialList.append(atrialDict)             #append one trial at a time to the final trialList

    return trialList

class StairHandler(_BaseTrialHandler):
    """Class to handle smoothly the selection of the next trial
    and report current values etc.
    Calls to nextTrial() will fetch the next object given to this
    handler, according to the method specified.

    See ``demo_trialHandler.py``

    The staircase will terminate when *nTrials* AND *nReversals* have been exceeded. If *stepSizes* was an array
    and has been exceeded before nTrials is exceeded then the staircase will continue
    to reverse

    """
    def __init__(self,
                 startVal,
                 nReversals=None,
                 stepSizes=4,  #dB stepsize
                 nTrials=0,
                 nUp=1,
                 nDown=3, #correct responses before stim goes down
                 extraInfo=None,
                 method = '2AFC',
                 stepType='db',
                 minVal=None,
                 maxVal=None,
                 originPath=None,
                 name=''):
        """
        :Parameters:

            startVal:
                The initial value for the staircase.

            nReversals:
                The minimum number of reversals permitted. If stepSizes is a list then there must
                also be enough reversals to satisfy this list.

            stepSizes:
                The size of steps as a single value or a list (or array). For a single value the step
                size is fixed. For an array or list the step size will progress to the next entry
                at each reversal.

            nTrials:
                The minimum number of trials to be conducted. If the staircase has not reached the
                required number of reversals then it will continue.

            nUp:
                The number of 'incorrect' (or 0) responses before the staircase level increases.

            nDown:
                The number of 'correct' (or 1) responses before the staircase level decreases.

            extraInfo:
                A dictionary (typically) that will be stored along with collected data using
                :func:`~psychopy.data.StairHandler.saveAsPickle` or
                :func:`~psychopy.data.StairHandler.saveAsText` methods.

            stepType:
                specifies whether each step will be a jump of the given size in
                'db', 'log' or 'lin' units ('lin' means this intensity will be added/subtracted)

            method:
                Not used and may be deprecated in future releases.

            stepType: *'db'*, 'lin', 'log'
                The type of steps that should be taken each time. 'lin' will simply add or subtract that
                amount each step, 'db' and 'log' will step by a certain number of decibels or log units
                (note that this will prevent your value ever reaching zero or less)

            minVal: *None*, or a number
                The smallest legal value for the staircase, which can be used to prevent it
                reaching impossible contrast values, for instance.

            maxVal: *None*, or a number
                The largest legal value for the staircase, which can be used to prevent it
                reaching impossible contrast values, for instance.

        """

        """
        trialList: a simple list (or flat array) of trials.

            """
        self.name=name
        self.startVal=startVal
        self.nReversals=nReversals
        self.nUp=nUp
        self.nDown=nDown
        self.extraInfo=extraInfo
        self.method=method
        self.stepType=stepType

        self.stepSizes=stepSizes
        if type(stepSizes) in [int, float]:
            self.stepSizeCurrent=stepSizes
            self._variableStep=False
        else:#list, tuple or array
            self.stepSizeCurrent=stepSizes[0]
            self.nReversals= max(len(stepSizes),self.nReversals)
            self._variableStep=True

        self.nTrials = nTrials#to terminate the nTrials must be exceeded and either
        self.finished=False
        self.thisTrialN = -1
        self.data = []
        self.intensities=[]
        self.reversalPoints = []
        self.reversalIntensities=[]
        self.currentDirection='start' #initially it goes down but on every step
        self.correctCounter=0  #correct since last stim change (minus are incorrect)
        self._nextIntensity=self.startVal
        self._warnUseOfNext=True
        self.minVal = minVal
        self.maxVal = maxVal

        #self.originPath and self.origin (the contents of the origin file)
        self.originPath, self.origin = self.getOriginPathAndFile(originPath)
        self._exp = None#the experiment handler that owns me!
    def __iter__(self):
        return self

    def addData(self, result, intensity=None):
        """Add a 1 or 0 to signify a correct/detected or incorrect/missed trial

        This is essential to advance the staircase to a new intensity level!

        Supplying an `intensity` value here indicates that you did not use the
        recommended intensity in your last trial and the staircase will
        replace its recorded value with the one you supplied here.
        """
        self.data.append(result)

        #if needed replace the existing intensity with this custom one
        if intensity!=None:
            self.intensities.pop()
            self.intensities.append(intensity)

        #increment the counter of correct scores
        if result==1:
            if len(self.data)>1 and self.data[-2]==result:
                #increment if on a run
                self.correctCounter+=1
            else:
                #or reset
                self.correctCounter = 1

        else:
            if  len(self.data)>1 and self.data[-2]==result:
                #increment if on a run
                self.correctCounter-=1
            else:
                #or reset
                self.correctCounter = -1

        if self.getExp()!=None:#update the experiment handler too
            self.getExp().addData("%s.result" %(self.name), result)
        self.calculateNextIntensity()

    def calculateNextIntensity(self):
        """based on current intensity, counter of correct responses and current direction"""

        if len(self.reversalIntensities)<1:
            #always using a 1-down, 1-up rule initially
            if self.data[-1]==1:    #last answer correct
                #got it right
                self._intensityDec()
                if self.currentDirection=='up':
                    reversal=True
                else:#direction is 'down' or 'start'
                    reversal=False
                    self.currentDirection='down'
            else:
                #got it wrong
                self._intensityInc()
                if self.currentDirection=='down':
                    reversal=True
                else:#direction is 'up' or 'start'
                    reversal=False
                #now:
                self.currentDirection='up'

        elif self.correctCounter >= self.nDown: #n right, time to go down!
            #make it harder
            self._intensityDec()
            if self.currentDirection!='down':
                reversal=True
            else:
                reversal=False
            self.currentDirection='down'

        elif self.correctCounter <= -self.nUp: #n wrong, time to go up!
            #make it easier
            self._intensityInc()
            #note current direction
            if self.currentDirection!='up':
                reversal=True
            else:
                reversal=False
            self.currentDirection='up'

        else:
            #same as previous trial
            reversal=False


        #add reversal info
        if reversal:
            self.reversalPoints.append(self.thisTrialN)
            self.reversalIntensities.append(self.intensities[-1])
        #test if we're done
        if len(self.reversalIntensities)>=self.nReversals and \
            len(self.intensities)>=self.nTrials:
                self.finished=True
        #new step size if necessary
        if reversal and self._variableStep and self.finished==False:
            if len(self.reversalIntensities) >= len(self.stepSizes):
                #we've gone beyond the list of step sizes so just use the last one
                self.stepSizeCurrent = self.stepSizes[-1]
            else:
                self.stepSizeCurrent = self.stepSizes[len(self.reversalIntensities)]


    def next(self):
        """Advances to next trial and returns it.
        Updates attributes; `thisTrial`, `thisTrialN` and `thisIndex`.

        If the trials have ended, calling this method will raise a StopIteration error.
        This can be handled with code such as::

            staircase = StairHandler(.......)
            for eachTrial in staircase:#automatically stops when done
                #do stuff

        or::

            staircase = StairHandler(.......)
            while True: #ie forever
                try:
                    thisTrial = staircase.next()
                except StopIteration:#we got a StopIteration error
                    break #break out of the forever loop
                #do stuff here for the trial

        """
        if self.finished==False:
            #update pointer for next trial
            self.thisTrialN+=1
            self.intensities.append(self._nextIntensity)
            return self._nextIntensity
        else:
            self._terminate()
    def _intensityInc(self):
        """increment the current intensity and reset counter"""
        if self.stepType=='db':
            self._nextIntensity *= 10.0**(self.stepSizeCurrent/20.0)
        elif self.stepType=='log':
            self._nextIntensity *= 10.0**self.stepSizeCurrent
        elif self.stepType=='lin':
            self._nextIntensity += self.stepSizeCurrent
        #check we haven't gone out of the legal range
        if (self._nextIntensity > self.maxVal) and self.maxVal is not None:
            self._nextIntensity = self.maxVal
        self.correctCounter =0

    def _intensityDec(self):
        """decrement the current intensity and reset counter"""
        if self.stepType=='db':
            self._nextIntensity /= 10.0**(self.stepSizeCurrent/20.0)
        if self.stepType=='log':
            self._nextIntensity /= 10.0**self.stepSizeCurrent
        elif self.stepType=='lin':
            self._nextIntensity -= self.stepSizeCurrent
        self.correctCounter =0
        #check we haven't gone out of the legal range
        if (self._nextIntensity < self.minVal) and self.minVal is not None:
            self._nextIntensity = self.minVal

    def saveAsText(self,fileName,
                   delim='\t',
                   matrixOnly=False,
                  ):
        """
        Write a text file with the data

        :Parameters:

            fileName: a string
                The name of the file, including path if needed. The extension
                `.dlm` will be added if not included.

            delim: a string
                the delimitter to be used (e.g. '\t' for tab-delimitted, ',' for csv files)

            matrixOnly: True/False
                If True, prevents the output of the `extraInfo` provided at initialisation.
        """

        if self.thisTrialN<1:
            logging.debug('StairHandler.saveAsText called but no trials completed. Nothing saved')
            return -1

        #create the file or print to stdout
        if fileName=='stdout':
            f = sys.stdout
        elif fileName[-4:] in ['.dlm','.DLM', '.csv','.CSV']:
            f= file(fileName,'w')
        else:
            if delim==',': f=file(fileName+'.csv','w')
            else: f=file(fileName+'.dlm','w')

        #write the data
        reversalStr = str(self.reversalIntensities)
        reversalStr = string.replace( reversalStr, ',', delim)
        reversalStr = string.replace( reversalStr, '[', '')
        reversalStr = string.replace( reversalStr, ']', '')
        f.write('\nreversalIntensities=\t%s\n' %reversalStr)

        reversalPts = str(self.reversalPoints)
        reversalPts = string.replace( reversalPts, ',', delim)
        reversalPts = string.replace( reversalPts, '[', '')
        reversalPts = string.replace( reversalPts, ']', '')
        f.write('reversalIndices=\t%s\n' %reversalPts)

        rawIntens = str(self.intensities)
        rawIntens = string.replace( rawIntens, ',', delim)
        rawIntens = string.replace( rawIntens, '[', '')
        rawIntens = string.replace( rawIntens, ']', '')
        f.write('\nintensities=\t%s\n' %rawIntens)

        responses = str(self.data)
        responses = string.replace( responses, ',', delim)
        responses = string.replace( responses, '[', '')
        responses = string.replace( responses, ']', '')
        f.write('responses=\t%s\n' %responses)

        #add self.extraInfo
        if (self.extraInfo != None) and not matrixOnly:
            strInfo = str(self.extraInfo)
            #dict begins and ends with {} - remove
            strInfo = strInfo[1:-1] #string.replace(strInfo, '{','');strInfo = string.replace(strInfo, '}','');
            strInfo = string.replace(strInfo, ': ', ':\n')#separate value from keyname
            strInfo = string.replace(strInfo, ',', '\n')#separate values from each other
            strInfo = string.replace(strInfo, 'array([ ', '')
            strInfo = string.replace(strInfo, '])', '')

            f.write('\n%s\n' %strInfo)

        f.write("\n")
        if f != sys.stdout:
            f.close()
            logging.info('saved data to %s' %f.name)

    def saveAsExcel(self,fileName, sheetName='data',
                   matrixOnly=False, appendFile=True,
                  ):
        """
        Save a summary data file in Excel OpenXML format workbook (:term:`xlsx`) for processing
        in most spreadsheet packages. This format is compatible with
        versions of Excel (2007 or greater) and and with OpenOffice (>=3.0).

        It has the advantage over the simpler text files (see :func:`TrialHandler.saveAsText()` )
        that data can be stored in multiple named sheets within the file. So you could have a single file
        named after your experiment and then have one worksheet for each participant. Or you could have
        one file for each participant and then multiple sheets for repeated sessions etc.

        The file extension `.xlsx` will be added if not given already.

        The file will contain a set of values specifying the staircase level ('intensity') at each
        reversal, a list of reversal indices (trial numbers), the raw staircase/intensity
        level on *every* trial and the corresponding responses of the participant on every trial.

        :Parameters:

            fileName: string
                the name of the file to create or append. Can include relative or absolute path

            sheetName: string
                the name of the worksheet within the file

            matrixOnly: True or False
                If set to True then only the data itself will be output (no additional info)

            appendFile: True or False
                If False any existing file with this name will be overwritten. If True then a new worksheet will be appended.
                If a worksheet already exists with that name a number will be added to make it unique.

        """

        if self.thisTrialN<1:
            logging.debug('StairHandler.saveAsExcel called but no trials completed. Nothing saved')
            return -1
        #NB this was based on the limited documentation (1 page wiki) for openpyxl v1.0
        if not haveOpenpyxl:
            raise ImportError, 'openpyxl is required for saving files in Excel (xlsx) format, but was not found.'
            return -1

        #import necessary subpackages - they are small so won't matter to do it here
        from openpyxl.workbook import Workbook
        from openpyxl.writer.excel import ExcelWriter
        from openpyxl.reader.excel import load_workbook

        if not fileName.endswith('.xlsx'): fileName+='.xlsx'
        #create or load the file
        if appendFile and os.path.isfile(fileName):
            wb = load_workbook(fileName)
            newWorkbook=False
        else:
            if not appendFile: #the file exists but we're not appending, so will be overwritten
                logging.warning('Data file, %s, will be overwritten' %fileName)
            wb = Workbook()#create new workbook
            wb.properties.creator='PsychoPy'+psychopy.__version__
            newWorkbook=True

        ew = ExcelWriter(workbook = wb)

        if newWorkbook:
            ws = wb.worksheets[0]
            ws.title=sheetName
        else:
            ws=wb.create_sheet()
            ws.title=sheetName

        #write the data
        #reversals data
        ws.cell('A1').value = 'Reversal Intensities'
        ws.cell('B1').value = 'Reversal Indices'
        for revN, revIntens in enumerate(self.reversalIntensities):
            ws.cell(_getExcelCellName(col=0,row=revN+1)).value = unicode(revIntens)
            ws.cell(_getExcelCellName(col=1,row=revN+1)).value = unicode(self.reversalPoints[revN])

        #trials data
        ws.cell('C1').value = 'All Intensities'
        ws.cell('D1').value = 'All Responses'
        for intenN, intensity in enumerate(self.intensities):
            ws.cell(_getExcelCellName(col=2,row=intenN+1)).value = unicode(intensity)
            ws.cell(_getExcelCellName(col=3,row=intenN+1)).value = unicode(self.data[intenN])

        #add self.extraInfo
        rowN = 0
        if (self.extraInfo != None) and not matrixOnly:
            ws.cell(_getExcelCellName(col=6,row=rowN)).value = 'extraInfo'; rowN+=1
            for key,val in self.extraInfo.items():
                ws.cell(_getExcelCellName(col=6,row=rowN)).value = unicode(key)+u':'
                ws.cell(_getExcelCellName(col=7,row=rowN)).value = unicode(val)
                rowN+=1

        ew.save(filename = fileName)
        logging.info('saved data to %s' %fileName)

    def saveAsPickle(self,fileName):
        """Basically just saves a copy of self (with data) to a pickle file.

        This can be reloaded if necess and further analyses carried out.
        """
        if self.thisTrialN<1:
            logging.debug('StairHandler.saveAsPickle called but no trials completed. Nothing saved')
            return -1
        #otherwise use default location
        f = open(fileName+'.psydat', "wb")
        cPickle.dump(self, f)
        f.close()
        logging.info('saved data to %s' %f.name)


class QuestHandler(StairHandler):
    """Class that implements the Quest algorithm for quick measurement of
    psychophysical thresholds.

    Uses Andrew Straw's `QUEST <http://www.visionegg.org/Quest>`_, which is a
    Python port of Denis Pelli's Matlab code.

    Measures threshold using a Weibull psychometric function. Currently, it is
    not possible to use a different psychometric function.

    Threshold 't' is measured on an abstract 'intensity' scale, which
    usually corresponds to log10 contrast.

    The Weibull psychometric function:

    p2=delta*gamma+(1-delta)*(1-(1-gamma)*exp(-10**(beta*(x2+xThreshold))))

    **Example**::

        # setup display/window
        ...
        # create stimulus
        stimulus = visual.RadialStim(win=win, tex='sinXsin', size=1, pos=[0,0], units='deg')
        ...
        # create staircase object
        # trying to find out the point where subject's response is 50/50
        # if wanted to do a 2AFC then the defaults for pThreshold and gamma are good
        staircase = data.QuestHandler(staircase._nextIntensity, 0.2, pThreshold=0.63, gamma=0.01,
                                  nTrials=20, minVal=0, maxVal=1)
        ...
        while thisContrast in staircase:
            # setup stimulus
            stimulus.setContrast(thisContrast)
            stimulus.draw()
            win.flip()
            core.wait(0.5)
            # get response
            ...
            # inform QUEST of the response, needed to calculate next level
            staircase.addData(thisResp)
        ...
        # can now access 1 of 3 suggested threshold levels
        staircase.mean()
        staircase.mode()
        staircase.quantile() #gets the median

    """
    def __init__(self,
                 startVal,
                 startValSd,
                 pThreshold=0.82,
                 nTrials=None,
                 stopInterval=None,
                 method='quantile',
                 stepType='log',
                 beta=3.5,
                 delta=0.01,
                 gamma=0.5,
                 grain=0.01,
                 range=None,
                 extraInfo=None,
                 minVal=None,
                 maxVal=None,
                 staircase=None,
                 originPath=None,
                 name=''):
        """
        Typical values for pThreshold are:
            * 0.82 which is equivalent to a 3 up 1 down standard staircase
            * 0.63 which is equivalent to a 1 up 1 down standard staircase (and might want gamma=0.01)

        The variable(s) nTrials and/or stopSd must be specified.

        `beta`, `delta`, and `gamma` are the parameters of the Weibull psychometric function.

        :Parameters:

            startVal:
                Prior threshold estimate or your initial guess threshold.

            startValSd:
                Standard deviation of your starting guess threshold. Be generous with the sd
                as QUEST will have trouble finding the true threshold if it's more than one sd
                from your initial guess.

            pThreshold
                Your threshold criterion expressed as probability of response==1. An intensity
                offset is introduced into the psychometric function so that the threshold (i.e.,
                the midpoint of the table) yields pThreshold..

            nTrials: *None* or a number
                The maximum number of trials to be conducted.

            stopInterval: *None* or a number
                The minimum 5-95% confidence interval required in the threshold estimate before stopping.
                If both this and nTrials is specified, whichever happens first will determine when
                Quest will stop.

            method: *'quantile'*, 'mean', 'mode'
                The method used to determine the next threshold to test. If you want to get a specific threshold
                level at the end of your staircasing, please use the quantile, mean, and mode methods directly.

            stepType: *'log'*, 'db', 'lin'
                The type of steps that should be taken each time. 'db' and 'log' will transform your intensity levels
                into decibels or log units and will move along the psychometric function with these values.

            beta: *3.5* or a number
                Controls the steepness of the psychometric function.

            delta: *0.01* or a number
                The fraction of trials on which the observer presses blindly.

            gamma: *0.5* or a number
                The fraction of trials that will generate response 1 when intensity=-Inf.

            grain: *0.01* or a number
                The quantization of the internal table.

            range: *None*, or a number
                The intensity difference between the largest and smallest intensity that the
                internal table can store. This interval will be centered on the initial guess
                tGuess. QUEST assumes that intensities outside of this range have zero prior
                probability (i.e., they are impossible).

            extraInfo:
                A dictionary (typically) that will be stored along with collected data using
                :func:`~psychopy.data.StairHandler.saveAsPickle` or
                :func:`~psychopy.data.StairHandler.saveAsText` methods.

            minVal: *None*, or a number
                The smallest legal value for the staircase, which can be used to prevent it
                reaching impossible contrast values, for instance.

            maxVal: *None*, or a number
                The largest legal value for the staircase, which can be used to prevent it
                reaching impossible contrast values, for instance.

            staircase: *None* or StairHandler
                Can supply a staircase object with intensities and results. Might be useful to
                give the quest algorithm more information if you have it. You can also call the
                importData function directly.

        """

        # Initialize using parent class first
        StairHandler.__init__(self, startVal, nTrials=nTrials, extraInfo=extraInfo, method=method,
                                stepType=stepType, minVal=minVal, maxVal=maxVal, name=name)

        # Setup additional values
        self.stopInterval = stopInterval

        # Transform startVal and startValSd based on stepType
        startVal = self._intensity2scale(startVal)
        startValSd = self._intensity2scale(startValSd)
        self._questNextIntensity = startVal

        # Create Quest object
        self._quest = QuestObject(startVal, startValSd, pThreshold, beta, delta, gamma, grain, range)

        # Import any old staircase data
        if staircase is not None:
            self.importData(staircase.intensities, staircase.data)
        #store the origin file and its path
        self.originPath, self.origin = self.getOriginPathAndFile(originPath)
        self._exp=None

    def addData(self, result, intensity=None):
        """Add a 1 or 0 to signify a correct/detected or incorrect/missed trial

        Supplying an `intensity` value here indicates that you did not use the
        recommended intensity in your last trial and the staircase will
        replace its recorded value with the one you supplied here.
        """
        # Process user supplied intensity
        if intensity is None:
            intensity = self._questNextIntensity
        else:
            intensity = self._intensity2scale(intensity)
            #update the intensity
            self.intensities.pop()#remove the one that had been auto-generated
            self.intensities.append(intensity)
        # Update quest
        self._quest.update(intensity, result)
        # Update other things
        self.data.append(result)
        if self.getExp()!=None:
            self.getExp().addData('response', result)
        self.calculateNextIntensity()

    def importData(self, intensities, results):
        """import some data which wasn't previously given to the quest algorithm"""
        # NOT SURE ABOUT CLASS TO USE FOR RAISING ERROR
        if len(intensities) != len(results):
            raise AttributeError, "length of intensities and results input must be the same"
        self.incTrials(len(intensities))
        for intensity, result in zip(intensities,results):
            try:
                self.next()
                self.addData(result, intensity)
            except StopIteration:   # would get a stop iteration if stopInterval set
                pass    # TODO: might want to check if nTrials is still good

    def calculateNextIntensity(self):
        """based on current intensity and counter of correct responses"""
        self._intensity()
        # Check we haven't gone out of the legal range
        if (self._nextIntensity > self.maxVal) and self.maxVal is not None:
            self._nextIntensity = self.maxVal
        elif (self._nextIntensity < self.minVal) and self.minVal is not None:
            self._nextIntensity = self.minVal
        self._questNextIntensity = self._intensity2scale(self._nextIntensity)
    def _intensity(self):
        """assigns the next intensity level"""
        if self.method == 'mean':
            self._questNextIntensity = self._quest.mean()
        elif self.method == 'mode':
            self._questNextIntensity = self._quest.mode()
        elif self.method == 'quantile':
            self._questNextIntensity = self._quest.quantile()
        # else: maybe raise an error
        self._nextIntensity = self._scale2intensity(self._questNextIntensity)

    def _intensity2scale(self, intensity):
        """returns the scaled intensity level based on value of self.stepType"""
        if self.stepType=='db':
            scaled_intensity = numpy.log10(intensity) * 20.0
        elif self.stepType=='log':
            scaled_intensity = numpy.log10(intensity)
        else:
            scaled_intensity = intensity
        return scaled_intensity

    def _scale2intensity(self, scaled_intensity):
        """returns the unscaled intensity level based on value of self.stepType"""
        if self.stepType=='db':
            intensity = 10.0**(scaled_intensity/20.0)
        elif self.stepType=='log':
            intensity = 10.0**scaled_intensity
        else:
            intensity = scaled_intensity
        return intensity

    def mean(self):
        """mean of Quest posterior pdf"""
        return self._scale2intensity(self._quest.mean())

    def sd(self):
        """standard deviation of Quest posterior pdf"""
        return self._scale2intensity(self._quest.sd())

    def mode(self):
        """mode of Quest posterior pdf"""
        return self._scale2intensity(self._quest.mode()[0])

    def quantile(self, p=None):
        """quantile of Quest posterior pdf"""
        return self._scale2intensity(self._quest.quantile(p))

    def confInterval(self, getDifference=False):
        """give the range of the 5-95% confidence interval"""
        interval = [self.quantile(0.05), self.quantile(0.95)]
        if getDifference:
            return abs(interval[0] - interval[1])
        else:
            return interval

    def incTrials(self, nNewTrials):
        """increase maximum number of trials
        Updates attribute: `nTrials`
        """
        self.nTrials += nNewTrials

    def simulate(self, tActual):
        """ returns a simulated user response to the next intensity level presented by Quest,
            need to supply the actual threshold level
        """
        # Current estimated intensity level
        if self.method == 'mean':
            tTest = self._quest.mean()
        elif self.method == 'mode':
            tTest = self._quest.mode()
        elif self.method == 'quantile':
            tTest = self._quest.quantile()
        return self._quest.simulate(tTest, tActual)

    def next(self):
        """Advances to next trial and returns it.
        Updates attributes; `thisTrial`, `thisTrialN`, `thisIndex`, `finished`, `intensities`

        If the trials have ended, calling this method will raise a StopIteration error.
        This can be handled with code such as::

            staircase = QuestHandler(.......)
            for eachTrial in staircase:#automatically stops when done
                #do stuff

        or::

            staircase = QuestHandler(.......)
            while True: #ie forever
                try:
                    thisTrial = staircase.next()
                except StopIteration:#we got a StopIteration error
                    break #break out of the forever loop
                #do stuff here for the trial
        """
        self._checkFinished()

        if self.finished==False:
            #update pointer for next trial
            self.thisTrialN+=1
            self.intensities.append(self._nextIntensity)
            return self._nextIntensity
        else:
            self._terminate()

    def _checkFinished(self):
        """checks if we are finished
        Updates attribute: `finished`
        """
        if self.nTrials is not None and len(self.intensities) >= self.nTrials:
            self.finished = True
        elif self.stopInterval is not None and self.confInterval(True) < self.stopInterval:
            self.finished = True
        else:
            self.finished = False


class MultiStairHandler(_BaseTrialHandler):
    def __init__(self, stairType='simple', method='random',
            conditions=None, nTrials=50, originPath=None, name=''):
        """A Handler to allow easy interleaved staircase procedures (simple or
        QUEST).

        Parameters for the staircases, as used by the relevant :class:`StairHandler` or
        :class:`QuestHandler` (e.g. the `startVal`, `minVal`, `maxVal`...)
        should be specified in the `conditions` list and may vary between
        each staircase. In particular, the conditions /must/ include the
        a `startVal` (because this is a required argument to the above handlers)
        a `label` to tag the staircase and a `startValSd` (only for QUEST
        staircases). Any parameters not specified in the conditions file
        will revert to the default for that individual handler.

        If you need to custom the behaviour further you may want to look at the
        recipe on :ref:`interleavedStairs`.

        :params:

            stairType: 'simple' or 'quest'
                Use a :class:`StairHandler` or :class:`QuestHandler`

            method: 'random' or 'sequential'
                The stairs are shuffled in each repeat but not randomised more than
                that (so you can't have 3 repeats of the same staircase in a row
                unless it's the only one still running)

            conditions: a list of dictionaries specifying conditions
                Can be used to control parameters for the different staicases.
                Can be imported from an Excel file using `psychopy.data.importTrialTypes`
                MUST include keys providing, 'startVal', 'label' and 'startValSd' (QUEST only).
                The 'label' will be used in data file saving so should be unique.
                See Example Usage below.

            nTrials=50
                Minimum trials to run (but may take more if the staircase hasn't
                also met its minimal reversals. See `~psychopy.data.StairHandler`

        Example usage::

            conditions=[
                {'label':'low', 'startVal': 0.1, 'ori':45},
                {'label':'high','startVal': 0.8, 'ori':45},
                {'label':'low', 'startVal': 0.1, 'ori':90},
                {'label':'high','startVal': 0.8, 'ori':90},
                ]
            stairs = MultiStairHandler(conditions=conditions, nTrials=50)

            for thisIntensity, thisCondition in stairs:
                thisOri = thisCondition['ori']

                #do something with thisIntensity and thisOri

                stairs.addData(correctIncorrect)#this is ESSENTIAL

            #save data as multiple formats
            stairs.saveDataAsExcel(fileName)#easy to browse
            stairs.saveAsPickle(fileName)#contains more info

        """
        self.name=name
        self.type=stairType
        self.method=method #'random' or 'sequential'
        self.conditions=conditions
        self.nTrials=nTrials
        self.finished=False
        self.totalTrials=0
        self._checkArguments()
        #create staircases
        self.staircases=[]#all staircases
        self.runningStaircases=[]#staircases that haven't finished yet
        self.thisPassRemaining=[]#staircases to run this pass
        self._createStairs()

        #fetch first staircase/value (without altering/advancing it)
        self._startNewPass()
        self.currentStaircase = self.thisPassRemaining[0]#take the first and remove it
        self._nextIntensity = self.currentStaircase._nextIntensity#gets updated by self.addData()
        #store the origin file and its path
        self.originPath, self.origin = self.getOriginPathAndFile(originPath)
        self._exp = None#the experiment handler that owns me!
    def _checkArguments(self):
        #did we get a conditions parameter, correctly formatted
        if type(self.conditions) not in [list]:
            logging.error('conditions parameter to MultiStairHandler should be a list, not a %s' %type(self.conditions))
            return
        c0=self.conditions[0]
        if type(c0)!=dict:
            logging.error('conditions to MultiStairHandler should be a list of python dictionaries' + \
                ', not a list of %ss' %type(c0))
        #did conditions contain the things we need?
        params = c0.keys()
        if self.type in ['simple','quest']:
            if 'startVal' not in params:
                logging.error('MultiStairHandler needs a param called `startVal` in conditions')
            if 'label' not in params:
                logging.error('MultiStairHandler needs a param called `label` in conditions')
            if 'startValSd' not in params and self.type=='quest':
                logging.error("MultiStairHandler('quest') needs a param called `startValSd` in conditions")
        else:
            logging.error("MultiStairHandler `stairType` should be 'simple' or 'quest', not '%s'" %self.type)
    def _createStairs(self):
        if self.type=='simple':
            defaults = {'nReversals':None, 'stepSizes':4, 'nTrials':self.nTrials,
                'nUp':1, 'nDown':3, 'extraInfo':None,
                'stepType':'db', 'minVal':None, 'maxVal':None}
        elif self.type=='quest':
            defaults = {'pThreshold':0.82, 'nTrials':self.nTrials, 'stopInterval':None,
                'method':'quantile', 'stepType':'log', 'beta':3.5, 'delta':0.01,
                'gamma':0.5, 'grain':0.01, 'range':None, 'extraInfo':None,
                'minVal':None, 'maxVal':None, 'staircase':None}

        for condition in self.conditions:
            startVal=condition['startVal']
            #fetch each params from conditions if possible
            for paramName in defaults:
                #get value for the parameter
                if paramName in condition.keys(): val=condition[paramName]
                else: val = defaults[paramName]
                #assign value to variable name
                exec('%s=%s' %(paramName, repr(val)))
            #then create actual staircase
            if self.type=='simple':
                thisStair = StairHandler(startVal, nReversals=nReversals,
                    stepSizes=stepSizes, nTrials=nTrials, nUp=nUp, nDown=nDown,
                    extraInfo=extraInfo,
                    stepType=stepType, minVal=minVal, maxVal=maxVal)
            elif self.type=='quest':
                thisStair = QuestHandler(startVal, startValSd=condition['startValSd'],
                    pThreshold=pThreshold, nTrials=nTrials, stopInterval=stopInterval,
                    method=method, stepType=stepType, beta=beta, delta=delta,
                    gamma=gamma, grain=grain, range=range, extraInfo=extraInfo,
                    minVal=minVal, maxVal=maxVal, staircase=staircase)
            thisStair.condition = condition#this isn't normally part of handler
            #and finally, add it to the list
            self.staircases.append(thisStair)
            self.runningStaircases.append(thisStair)
    def __iter__(self):
        return self
    def next(self):
        """Advances to next trial and returns it.

        This can be handled with code such as::

            staircase = MultiStairHandler(.......)
            for eachTrial in staircase:#automatically stops when done
                #do stuff here for the trial

        or::

            staircase = MultiStairHandler(.......)
            while True: #ie forever
                try:
                    thisTrial = staircase.next()
                except StopIteration:#we got a StopIteration error
                    break #break out of the forever loop
                #do stuff here for the trial

        """
        #create a new set for this pass if needed
        if not hasattr(self, 'thisPassRemaining') or self.thisPassRemaining==[]:
            if len(self.runningStaircases)>0:
                self._startNewPass()
            else:
                self.finished=True
                raise StopIteration
        #fetch next staircase/value
        self.currentStaircase = self.thisPassRemaining.pop(0)#take the first and remove it
        self._nextIntensity = self.currentStaircase._nextIntensity#gets updated by self.addData()
        #return value
        if self.finished==False:
            return self._nextIntensity, self.currentStaircase.condition
        else:
            raise StopIteration

    def _startNewPass(self):
        """Create a new iteration of the running staircases for this pass.

        This is not normally needed byt he user - it gets called at __init__
        and every time that next() runs out of trials for this pass.
        """
        self.thisPassRemaining = copy.copy(self.runningStaircases)
        if self.method=='random': numpy.random.shuffle(self.thisPassRemaining)
    def addData(self, result, intensity=None):
        """Add a 1 or 0 to signify a correct/detected or incorrect/missed trial

        This is essential to advance the staircase to a new intensity level!
        """
        self.currentStaircase.addData(result)
        if self.getExp()!=None:#update the experiment handler too
            self.getExp().addData('response', result)
        try:
            self.currentStaircase.next()
        except:
            self.runningStaircases.remove(self.currentStaircase)
        self.totalTrials+=1
    def saveAsPickle(self, fileName):
        """Saves a copy of self (with data) to a pickle file.

        This can be reloaded later and further analyses carried out.
        """
        if self.totalTrials<1:
            logging.debug('StairHandler.saveAsPickle called but no trials completed. Nothing saved')
            return -1
        #otherwise use default location
        f = open(fileName+'.psydat', "wb")
        cPickle.dump(self, f)
        f.close()
        logging.info('saved data to %s' %f.name)
    def saveAsExcel(self, fileName, matrixOnly=False, appendFile=False):
        """
        Save a summary data file in Excel OpenXML format workbook (:term:`xlsx`) for processing
        in most spreadsheet packages. This format is compatible with
        versions of Excel (2007 or greater) and and with OpenOffice (>=3.0).

        It has the advantage over the simpler text files (see :func:`TrialHandler.saveAsText()` )
        that the data from each staircase will be save in the same file, with
        the sheet name coming from the 'label' given in the dictionary of
        conditions during initialisation of the Handler.

        The file extension `.xlsx` will be added if not given already.

        The file will contain a set of values specifying the staircase level ('intensity') at each
        reversal, a list of reversal indices (trial numbers), the raw staircase/intensity
        level on *every* trial and the corresponding responses of the participant on every trial.

        :Parameters:

            fileName: string
                the name of the file to create or append. Can include relative or absolute path

            matrixOnly: True or False
                If set to True then only the data itself will be output (no additional info)

            appendFile: True or False
                If False any existing file with this name will be overwritten. If True then a new worksheet will be appended.
                If a worksheet already exists with that name a number will be added to make it unique.

        """
        if self.totalTrials<1:
            logging.debug('StairHandler.saveAsExcel called but no trials completed. Nothing saved')
            return -1
        for stairN, thisStair in enumerate(self.staircases):
            if stairN==0: append=appendFile
            else: append=True
            #make a filename
            label = thisStair.condition['label']
            thisStair.saveAsExcel(fileName=fileName, sheetName=label,
                matrixOnly=matrixOnly, appendFile=append)
    def saveAsText(self,fileName,
                   delim='\t',
                   matrixOnly=False):
        """
        Write out text files with the data.

        For MultiStairHandler this will output one file for each staircase
        that was run, with _label added to the fileName that you specify above
        (label comes from the condition dictionary you specified when you
        created the Handler).

        :Parameters:

            fileName: a string
                The name of the file, including path if needed. The extension
                `.dlm` will be added if not included.

            delim: a string
                the delimitter to be used (e.g. '\t' for tab-delimitted, ',' for csv files)

            matrixOnly: True/False
                If True, prevents the output of the `extraInfo` provided at initialisation.
        """
        if self.totalTrials<1:
            logging.debug('StairHandler.saveAsText called but no trials completed. Nothing saved')
            return -1
        for stairN, thisStair in enumerate(self.staircases):
            #make a filename
            label = thisStair.condition['label']
            thisFileName = fileName+"_"+label
            thisStair.saveAsText(fileName=thisFileName, delim=delim,
                matrixOnly=matrixOnly)
    def printAsText(self,
                   delim='\t',
                   matrixOnly=False):
        """
        Write the data to the standard output stream

        :Parameters:

            delim: a string
                the delimitter to be used (e.g. '\t' for tab-delimitted, ',' for csv files)

            matrixOnly: True/False
                If True, prevents the output of the `extraInfo` provided at initialisation.
        """
        nStairs=len(self.staircases)
        for stairN, thisStair in enumerate(self.staircases):
            if stairN<(nStairs-1): thisMatrixOnly=True #never print info for first files
            else: thisMatrixOnly = matrixOnly
            #make a filename
            label = thisStair.condition['label']
            print "\n%s:" %label
            thisStair.saveAsText(fileName='stdout', delim=delim,
                matrixOnly=thisMatrixOnly)

class DataHandler(dict):
    """For handling data (used by TrialHandler, principally, rather than
    by users directly)

    Numeric data are stored as numpy masked arrays where the mask is set True for missing entries.
    When any non-numeric data (string, list or array) get inserted using DataHandler.add(val) the array
    is converted to a standard (not masked) numpy array with dtype='O' and where missing entries have
    value="--"

    Attributes:
        - ['key']=data arrays containing values for that key
            (e.g. data['accuracy']=...)
        - dataShape=shape of data (x,y,...z,nReps)
        - dataTypes=list of keys as strings

    """
    def __init__(self, dataTypes=None, trials=None, dataShape=None):
        self.trials=trials
        self.dataTypes=[]#names will be added during addDataType
        self.isNumeric={}
        #if given dataShape use it - otherwise guess!
        if dataShape: self.dataShape=dataShape
        elif self.trials:
            self.dataShape=list(numpy.asarray(trials.trialList,'O').shape)
            self.dataShape.append(trials.nReps)

        #initialise arrays now if poss
        if dataTypes and self.dataShape:
            for thisType in dataTypes:
                self.addDataType(thisType)

    def addDataType(self, names, shape=None):
        """Add a new key to the data dictionary of
        particular shape if specified (otherwise the
        shape of the trial matrix in the trial handler.
        Data are initialised to be zero everywhere.
        Not needed by user: appropriate types will be added
        during initialisation and as each xtra type is needed.
        """
        if not shape: shape = self.dataShape
        if not isinstance(names,basestring):
            #recursively call this function until we have a string
            for thisName in names: self.addDataType(thisName)
        else:
            #create the appropriate array in the dict
            #initially use numpy masked array of floats with mask=True for missing vals
            #convert to a numpy array with dtype='O' if non-numeric data given
            #NB don't use masked array with dytpe='O' together -they don't unpickle
            self[names]=numpy.ma.zeros(shape,'f')#masked array of floats
            self[names].mask=True
            #add the name to the list
            self.dataTypes.append(names)
            self.isNumeric[names]=True#until we need otherwise
    def add(self, thisType, value, position=None):
        """Add data to an existing data type
        (and add a new one if necess)
        """
        if not self.has_key(thisType):
            self.addDataType(thisType)
        if position==None:
            #'ran' is always the first thing to update
            if thisType=='ran':
                repN = sum(self['ran'][self.trials.thisIndex])
            else:
                repN = sum(self['ran'][self.trials.thisIndex])-1#because it has already been updated
            #make a list where 1st digit is trial number
            position= [self.trials.thisIndex]
            position.append(repN)

        #check whether data falls within bounds
        posArr = numpy.asarray(position)
        shapeArr = numpy.asarray(self.dataShape)
        if not numpy.alltrue(posArr<shapeArr):
            #array isn't big enough
            logging.warning('need a bigger array for:'+thisType)
            self[thisType]=misc.extendArr(self[thisType],posArr)#not implemented yet!
        #check for ndarrays with more than one value and for non-numeric data
        if self.isNumeric[thisType] and \
            ((type(value)==numpy.ndarray and len(value)>1) or (type(value) not in [float, int])):
                self._convertToObjectArray(thisType)
        #insert the value
        self[thisType][position[0],position[1]]=value
    def _convertToObjectArray(self, thisType):
        """Convert this datatype from masked numeric array to unmasked object array
        """
        dat = self[thisType]
        self[thisType] = numpy.array(dat.data, dtype='O')#create an array of Object type
        #masked vals should be "--", others keep data
        self[thisType] = numpy.where(dat.mask, '--',dat).astype('O')#we have to repeat forcing to 'O' or text gets truncated to 4chars
        self.isNumeric[thisType]=False

class FitFunction:
    """Deprecated - use the specific functions; FitWeibull, FitLogistic...
    """

    def __init__(self, fnName, xx, yy, sems=1.0, guess=None, display=1,
                 expectedMin=0.5):
        self.fnName = fnName
        self.xx = numpy.asarray(xx)
        self.yy = numpy.asarray(yy)
        self.sems = numpy.asarray(sems)
        self.params = guess
        self.display=display
        # for holding error calculations:
        self.ssq=0
        self.rms=0
        self.chi=0

        if fnName[-4:] in ['2AFC', 'TAFC']:
            self.expectedMin = 0.5
        elif fnName[-2:] =='YN':
            self.expectedMin=0.0
        else:
            self.expectedMin=expectedMin

        #do the calculations:
        self._doFit()

    def _doFit(self):
        #get some useful variables to help choose starting fit vals
        xMin = min(self.xx); xMax = max(self.xx)
        xRange=xMax-xMin; xMean= (xMax+xMin)/2.0
        if self.fnName in ['weibullTAFC','weibullYN']:
            if self.params==None: guess=[xMean, xRange/5.0]
            else: guess= numpy.asarray(self.params,'d')
        elif self.fnName in ['cumNorm','erf']:
            if self.params==None: guess=[xMean, xRange/5.0]#c50, xScale (slope)
            else: guess= numpy.asarray(self.params,'d')
        elif self.fnName in ['logisticTAFC','logistYN', 'logistic']:
            if self.params==None: guess=[xMin, 5.0/xRange]#x0, xRate
            else: guess= numpy.asarray(self.params,'d')
        elif self.fnName in ['nakaRush', 'nakaRushton', 'NR']:
            if self.params==None: guess=[xMean, 2.0]#x50, expon
            else: guess= numpy.asarray(self.params,'d')

        self.params = optimize.fmin_cg(self._getErr, guess, None, (self.xx,self.yy,self.sems),disp=self.display)
        self.ssq = self._getErr(self.params, self.xx, self.yy, 1.0)
        self.chi = self._getErr(self.params, self.xx, self.yy, self.sems)
        self.rms = self.ssq/len(self.xx)

    def _getErr(self, params, xx,yy,sems):
        mod = self.eval(xx, params)
        err = sum((yy-mod)**2/sems)
        return err

    def eval(self, xx=None, params=None):
        if xx==None: xx=self.xx
        if params==None: params=self.params
        if self.fnName in ['weibullTAFC', 'weibull2AFC']:
            alpha = params[0];
            if alpha<=0: alpha=0.001
            beta = params[1]
            xx = numpy.asarray(xx)
            yy =  1.0 - 0.5*numpy.exp( - (xx/alpha)**beta )
        elif self.fnName == 'weibullYN':
            alpha = params[0];
            if alpha<=0: alpha=0.001
            beta = params[1]
            xx = numpy.asarray(xx)
            yy =  1.0 - numpy.exp( - (xx/alpha)**beta )
        elif self.fnName in ['nakaRush', 'nakaRushton', 'NR']:
            c50 = params[0]
            if c50<=0: c50=0.001
            n = params[1]
            if n<=0: n=0.001
            xx = numpy.asarray(xx)
            yy = rMax*(xx**n/(xx**n+c50**n))
        elif self.fnName in [ 'erf', 'cumNorm']:
            xShift = params[0]
            xScale = params[1]
            if xScale<=0: xScale=0.001
            xx = numpy.asarray(xx)
            yy = special.erf(xx*xScale - xShift)*0.5+0.5#numpy.special.erf() goes from -1:1
        elif self.fnName in [ 'logisticYN', 'logistYN']:
            x0 = params[0]
            xRate = params[1]
            if xRate<=0: xRate=0.001
            xx = numpy.asarray(xx)
            yy = 1.0/(1+(1.0/x0-1)*numpy.exp(-xRate*xx))
        return yy

    def inverse(self, yy, params=None):
        """Returns fitted xx for any given yy value(s).

        If params is specified this will override the current model params.
        """
        yy = numpy.asarray(yy)
        if params==None: params=self.params
        if self.fnName== 'weibullTAFC':
            alpha = params[0]
            beta = params[1]
            xx = alpha * (-numpy.log(2.0 * (1.0-yy))) **(1.0/beta)
        elif self.fnName== 'weibullYN':
            alpha = params[0]
            beta = params[1]
            xx = alpha * (-numpy.log(1.0-yy))**(1.0/beta)
        elif self.fnName in [ 'erf', 'cumNorm']:
            xShift = params[0]
            xScale = params[1]
            xx = (special.erfinv(yy*2.0-1.0)+xShift)/xScale
        elif self.fnName in [ 'logisticYN', 'logistYN']:
            x0 = params[0]
            xRate = params[1]
            xx = -numpy.log( (1/yy-1)/(1/x0-1) )/xRate
        elif self.fnName in ['nakaRush', 'nakaRushton', 'NR']:
            c50 = params[0]
            n = params[1]
            xx = c50/(1/yy-1)
        return xx

class _baseFunctionFit:
    """Not needed by most users except as a superclass for developping your own functions

    You must overide the eval and inverse methods and a good idea to overide the _initialGuess
    method aswell.
    """

    def __init__(self, xx, yy, sems=1.0, guess=None, display=1,
                 expectedMin=0.5):
        self.xx = numpy.asarray(xx)
        self.yy = numpy.asarray(yy)
        self.sems = numpy.asarray(sems)
        self.expectedMin = expectedMin
        self.display=display
        # for holding error calculations:
        self.ssq=0
        self.rms=0
        self.chi=0
        #initialise parameters
        if guess==None:
            self.params = self._initialGuess()
        else:
            self.params = guess

        #do the calculations:
        self._doFit()

    def _doFit(self):
        #get some useful variables to help choose starting fit vals
        self.params = optimize.fmin_powell(self._getErr, self.params, (self.xx,self.yy,self.sems),disp=self.display)
#        self.params = optimize.fmin_bfgs(self._getErr, self.params, None, (self.xx,self.yy,self.sems),disp=self.display)
        self.ssq = self._getErr(self.params, self.xx, self.yy, 1.0)
        self.chi = self._getErr(self.params, self.xx, self.yy, self.sems)
        self.rms = self.ssq/len(self.xx)

    def _initialGuess(self):
        xMin = min(self.xx); xMax = max(self.xx)
        xRange=xMax-xMin; xMean= (xMax+xMin)/2.0
        guess=[xMean, xRange/5.0]
        return guess

    def _getErr(self, params, xx,yy,sems):
        mod = self.eval(xx, params)
        err = sum((yy-mod)**2/sems)
        return err

    def eval(self, xx=None, params=None):
        """Returns fitted yy for any given xx value(s).
        Uses the original xx values (from which fit was calculated)
        if none given.

        If params is specified this will override the current model params."""
        yy=xx
        return yy

    def inverse(self, yy, params=None):
        """Returns fitted xx for any given yy value(s).

        If params is specified this will override the current model params.
        """
        #define the inverse for your function here
        xx=yy
        return xx


class FitWeibull(_baseFunctionFit):
    """Fit a Weibull function (either 2AFC or YN)
    of the form::

        y = chance + (1.0-chance)*(1-exp( -(xx/alpha)**(beta) ))

    and with inverse::

        x = alpha * (-log((1.0-y)/(1-chance)))**(1.0/beta)

    After fitting the function you can evaluate an array of x-values
    with ``fit.eval(x)``, retrieve the inverse of the function with
    ``fit.inverse(y)`` or retrieve the parameters from ``fit.params``
    (a list with ``[alpha, beta]``)"""
    def eval(self, xx=None, params=None):
        if params==None:  params=self.params #so the user can set params for this particular eval
        alpha = params[0];
        if alpha<=0: alpha=0.001
        beta = params[1]
        xx = numpy.asarray(xx)
        yy =  self.expectedMin + (1.0-self.expectedMin)*(1-numpy.exp( -(xx/alpha)**(beta) ))
        return yy
    def inverse(self, yy, params=None):
        if params==None: params=self.params #so the user can set params for this particular inv
        alpha = params[0]
        beta = params[1]
        xx = alpha * (-numpy.log((1.0-yy)/(1-self.expectedMin))) **(1.0/beta)
        return xx
class FitNakaRushton(_baseFunctionFit):
    """Fit a Naka-Rushton function
    of the form::

        yy = rMin + (rMax-rMin) * xx**n/(xx**n+c50**n)

    After fitting the function you can evaluate an array of x-values
    with ``fit.eval(x)``, retrieve the inverse of the function with
    ``fit.inverse(y)`` or retrieve the parameters from ``fit.params``
    (a list with ``[rMin, rMax, c50, n]``)

    Note that this differs from most of the other functions in
    not using a value for the expected minimum. Rather, it fits this
    as one of the parameters of the model."""
    def __init__(self, xx, yy, sems=1.0, guess=None, display=1):
        self.xx = numpy.asarray(xx)
        self.yy = numpy.asarray(yy)
        self.sems = numpy.asarray(sems)
        self.display=display
        # for holding error calculations:
        self.ssq=0
        self.rms=0
        self.chi=0
        #initialise parameters
        if guess==None:
            self.params = self._initialGuess()
        else:
            self.params = guess

        #do the calculations:
        self._doFit()
    def _initialGuess(self):
        xMin = min(self.xx); xMax = max(self.xx)
        xRange=xMax-xMin; xMean= (xMax+xMin)/2.0
        guess=[xMean, 2.0, min(self.yy), max(self.yy)-min(self.yy)]
        return guess
    def eval(self, xx=None, params=None):
        if params==None:  params=self.params #so the user can set params for this particular eval
        c50 = params[0]
        n = params[1]
        rMin = params[2]
        rMax = params[3]
        #all params should be >0
        if c50<=0: c50=0.001
        if n<=0: n=0.001
        if rMax<=0: n=0.001
        if rMin<=0: n=0.001

        xx = numpy.asarray(xx)
        yy = rMin + (rMax-rMin)*(xx**n/(xx**n+c50**n))
        #yy = (xx**n/(xx**n+c50**n))
        return yy

    def inverse(self, yy, params=None):
        if params==None: params=self.params #so the user can set params for this particular inv
        yy=numpy.asarray(yy)
        c50 = params[0]
        n = params[1]
        rMin = params[2]
        rMax = params[3]

        yScaled = (yy-rMin)/(rMax-rMin) #remove baseline and scale
        xx = (yScaled*c50**n/(1-yScaled))**(1/n)
        return xx

class FitLogistic(_baseFunctionFit):
    """Fit a Logistic function (either 2AFC or YN)
    of the form::

        y = chance + (1-chance)/(1+exp((PSE-xx)*JND))

    and with inverse::

        x = PSE - log((1-chance)/(yy-chance) - 1)/JND

    After fitting the function you can evaluate an array of x-values
    with ``fit.eval(x)``, retrieve the inverse of the function with
    ``fit.inverse(y)`` or retrieve the parameters from ``fit.params``
    (a list with ``[PSE, JND]``)
    """
    def eval(self, xx=None, params=None):
        if params==None:  params=self.params #so the user can set params for this particular eval
        PSE = params[0]
        JND = params[1]
        chance = self.expectedMin
        xx = numpy.asarray(xx)
        yy = chance + (1-chance)/(1+numpy.exp((PSE-xx)*JND))
        return yy
    def inverse(self, yy, params=None):
        if params==None: params=self.params #so the user can set params for this particular inv
        PSE = params[0]
        JND = params[1]
        chance = self.expectedMin
        yy = numpy.asarray(yy)
        xx = PSE - numpy.log((1-chance)/(yy-chance) - 1)/JND
        return xx

class FitCumNormal(_baseFunctionFit):
    """Fit a Cumulative Normal function (aka error function or erf)
    of the form::

        y = chance + (1-chance)*(special.erf(xx*xScale - xShift)/2.0+0.5)

    and with inverse::

        x = (erfinv((yy-chance)/(1-chance)*2.0-1)+xShift)/xScale

    After fitting the function you can evaluate an array of x-values
    with fit.eval(x), retrieve the inverse of the function with
    fit.inverse(y) or retrieve the parameters from fit.params
    (a list with [centre, sd] for the Gaussian distribution forming the cumulative)

    NB: Prior to version 1.74 the parameters had different meaning, relating
    to xShift and slope of the function (similar to 1/sd). Although that is more in
    with the parameters for the Weibull fit, for instance, it is less in keeping
    with standard expectations of normal (Gaussian distributions) so in version
    1.74.00 the parameters became the [centre,sd] of the normal distribution.

    """
    def eval(self, xx=None, params=None):
        if params==None:  params=self.params #so the user can set params for this particular eval
        xShift = params[0]
        sd = params[1]
        chance = self.expectedMin
        #if xScale<=0: xScale=0.001
        xx = numpy.asarray(xx)
        yy = chance + (1-chance)*(special.erf((xx-xShift)/sd)/2.0+0.5)#NB numpy.special.erf() goes from -1:1
        return yy
    def inverse(self, yy, params=None):
        if params==None: params=self.params #so the user can set params for this particular inv
        xShift = params[0]
        sd = params[1]
        chance = self.expectedMin
        #xx = (special.erfinv((yy-chance)/(1-chance)*2.0-1)+xShift)/xScale#NB numpy.special.erfinv() goes from -1:1
        xx = xShift+sd*special.erfinv(( (yy-chance)/(1-chance) - 0.5 )*2)
        return xx



########################## End psychopy.data classes ##########################

def bootStraps(dat, n=1):
    """Create a list of n bootstrapped resamples of the data

    SLOW IMPLEMENTATION (Python for-loop)

    Usage:
        ``out = bootStraps(dat, n=1)``

    Where:
        dat
            an NxM or 1xN array (each row is a different condition, each column is a different trial)
        n
            number of bootstrapped resamples to create

        out
            - dim[0]=conditions
            - dim[1]=trials
            - dim[2]=resamples
    """
    dat = numpy.asarray(dat)
    if len(dat.shape)==1: #have presumably been given a series of data for one stimulus
        dat=numpy.array([dat])#adds a dimension (arraynow has shape (1,Ntrials))

    nTrials = dat.shape[1]
    #initialise a matrix to store output
    resamples = numpy.zeros(dat.shape+(n,), dat.dtype)
    for stimulusN in range(dat.shape[0]):
        thisStim = dat[stimulusN,:]#fetch data for this stimulus
        for sampleN in range(n):
            indices = numpy.floor(nTrials*numpy.random.rand(nTrials)).astype('i')
            resamples[stimulusN,:,sampleN] = numpy.take(thisStim, indices)
    return resamples

def functionFromStaircase(intensities, responses, bins = 10):
    """Create a psychometric function by binning data from a staircase procedure

    usage::

        [intensity, meanCorrect, n] = functionFromStaircase(intensities, responses, bins)

    where:
            intensities
                are a list of intensities to be binned

            responses
                are a list of 0,1 each corresponding to the equivalent intensity value

            bins
                can be an integer (giving that number of bins) or 'unique' (where each bin is made from ALL data for exactly one intensity value)

            intensity
                is the center of an intensity bin

            meanCorrect
                is mean % correct in that bin

            n
                is number of responses contributing to that mean
    """
    #convert to arrays
    try:#concatenate if multidimensional
        intensities = numpy.concatenate(intensities)
        responses = numpy.concatenate(responses)
    except:
        intensities = numpy.array(intensities)
        responses = numpy.array(responses)

    #sort the responses
    sort_ii = numpy.argsort(intensities)
    sortedInten = numpy.take(intensities, sort_ii)
    sortedResp = numpy.take(responses, sort_ii)

    binnedResp=[]; binnedInten=[]; nPoints = []
    if bins=='unique':
        intensities = numpy.round(intensities, decimals=8)
        uniqueIntens=numpy.unique(intensities)
        for thisInten in uniqueIntens:
            theseResps = responses[intensities==thisInten]
            binnedInten.append(thisInten)
            binnedResp.append(numpy.mean(theseResps))
            nPoints.append(len(theseResps))
    else:
        pointsPerBin = len(intensities)/float(bins)
        for binN in range(bins):
            thisResp = sortedResp[int(round(binN*pointsPerBin)) : int(round((binN+1)*pointsPerBin))]
            thisInten = sortedInten[int(round(binN*pointsPerBin)) : int(round((binN+1)*pointsPerBin))]

            binnedResp.append( numpy.mean(thisResp))
            binnedInten.append( numpy.mean(thisInten))
            nPoints.append( len(thisInten) )

    return binnedInten, binnedResp, nPoints

def getDateStr(format="%Y_%b_%d_%H%M"):
    """Uses ``time.strftime()``_ to generate a string of the form
    2012_Apr_19_1531 for 19th April 3.31pm, 2012.
    This is often useful appended to data filenames to provide unique names.
    To include the year: getDateStr(format="%Y_%b_%d_%H%M") returns '2011_Mar_16_1307'
    depending on locale, can have unicode chars in month names, so utf_8_decode them
    For date in the format of the current localization, do:
        data.getDateStr(format=locale.nl_langinfo(locale.D_T_FMT))
    """
    now = time.strftime(format, time.localtime())
    try:
        now_dec = codecs.utf_8_decode(now)[0]
    except UnicodeDecodeError:
        now_dec = time.strftime("%Y_%m_%d_%H%M", time.localtime())  # '2011_03_16_1307'

    return now_dec

def isValidVariableName(name):
    """Checks whether a certain string could be used as a valid variable.

    Usage::

        OK, msg = isValidVariableName(name)

    >>> isValidVariableName('name')
    (True, '')
    >>> isValidVariableName('0name')
    (False, 'Variables cannot begin with numeric character')
    >>> isValidVariableName('first second')
    (False, 'Variables cannot contain punctuation or spaces')
    >>> isValidVariableName('')
    (False, "Variables cannot be missing, None, or ''")
    >>> isValidVariableName(None)
    (False, "Variables cannot be missing, None, or ''")
    >>> isValidVariableName(23)
    (False, "Variables must be string-like")
    >>> isValidVariableName('a_b_c')
    (True, '')
    """
    if not name:
        return False, "Variables cannot be missing, None, or ''"
    if not type(name) in [str, unicode, numpy.string_, numpy.unicode_]:
        return False, "Variables must be string-like"
    try:
        name=str(name)#convert from unicode if possible
    except:
        if type(name) in [unicode, numpy.unicode_]:
            raise AttributeError, "name %s (type %s) contains non-ASCII characters (e.g. accents)" % (name, type(name))
        else:
            raise AttributeError, "name %s (type %s) could not be converted to a string" % (name, type(name))

    if name[0].isdigit():
        return False, "Variables cannot begin with numeric character"
    if _nonalphanumeric_re.search(name):
        return False, "Variables cannot contain punctuation or spaces"
    return True, ""

def _getExcelCellName(col, row):
    """Returns the excel cell name for a row and column (zero-indexed)

    >>> _getExcelCellName(0,0)
    'A1'
    >>> _getExcelCellName(2,1)
    'C2'
    """
    return "%s%i" %(get_column_letter(col+1), row+1)#BEWARE - openpyxl uses indexing at 1, to fit with Excel

